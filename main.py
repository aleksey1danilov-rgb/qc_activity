from fastapi import FastAPI, Request, Depends, HTTPException, Form, Query
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, Response
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session
from sqlalchemy import desc, func, case, and_, or_
from datetime import datetime, date, timedelta
from typing import Optional, List, Dict, Any
import calendar
import os
import sys
import hashlib
import hmac
import json
from io import BytesIO

from dotenv import load_dotenv
load_dotenv()

from sqlalchemy import inspect, text

# Для экспорта в Excel
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ Библиотеки для Excel не установлены. Экспорт будет недоступен.")
    print("   Установите: pip install pandas openpyxl")

from database import (
    get_db, init_db, Project, Metric, MetricBlock, Operator, 
    Evaluation, EvaluationMetric, EvaluationBlockResult,  
    User, UserRole, ProjectManager, Ticket, TicketHistory, TicketStatus, TicketPriority, TicketTopic,
    calculate_evaluation_scores, Topic, update_operator_avg_quality
)
from auth import router as auth_router, get_current_user, get_optional_user
from reports import (
    get_operator_stats, get_project_stats, 
    get_weekly_dynamics, get_monthly_dynamics
)

app = FastAPI(debug=True, title="Оценка звонков")

# Подключаем роутер авторизации
app.include_router(auth_router)

# Шаблоны
templates = Jinja2Templates(
    directory="templates", 
    auto_reload=True,
    cache_size=0
)

# Статика
app.mount("/static", StaticFiles(directory="static"), name="static")

# ============== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ==============

def require_role(allowed_roles):
    """Декоратор для проверки роли"""
    def role_checker(current_user: User = Depends(get_current_user)):
        if current_user.role not in allowed_roles:
            raise HTTPException(status_code=403, detail="Недостаточно прав")
        return current_user
    return role_checker

def calculate_sla_deadline(start_date: date, days: int = 2) -> date:
    """Рассчитывает дедлайн с учетом рабочих дней"""
    current = start_date
    workdays_count = 0
    
    while workdays_count < days:
        current += timedelta(days=1)
        # Пропускаем выходные (суббота и воскресенье)
        if current.weekday() < 5:  # 0-4 это понедельник-пятница
            workdays_count += 1
    
    return current

def get_active_tickets_count(db: Session) -> int:
    """Получение количества активных тикетов (новые + в работе)"""
    return db.query(Ticket).filter(
        Ticket.status.in_([TicketStatus.NEW.value, TicketStatus.IN_PROGRESS.value])
    ).count()

def get_user_projects(user: User, db: Session) -> List[int]:
    """Получение списка проектов, доступных пользователю"""
    if user.role == 'admin':
        # Админ видит все проекты
        projects = db.query(Project.id).filter(Project.is_active == True).all()
        return [p[0] for p in projects]
    elif user.role == 'manager':
        # Менеджер видит свои проекты
        return [pm.project_id for pm in user.manager_links]
    elif user.role == 'controller':
        # Контролер видит все проекты (можно изменить при необходимости)
        projects = db.query(Project.id).filter(Project.is_active == True).all()
        return [p[0] for p in projects]
    else:
        return []

def save_virtual_workbook(workbook):
    """Сохранение workbook в байтовый поток"""
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()

# ============== ИНИЦИАЛИЗАЦИЯ ==============

@app.on_event("startup")
def startup_event():
    """Создание таблиц и проверка структуры БД при запуске"""
    print("\n" + "="*60)
    print("🚀 ЗАПУСК ПРИЛОЖЕНИЯ")
    print("="*60)
    
    # Создаем таблицы, если их нет
    init_db()
    print("✅ Таблицы созданы/проверены")
    
    # ============== МИГРАЦИЯ ТАБЛИЦЫ METRICS ==============
    print("\n📦 ПРОВЕРКА СТРУКТУРЫ ТАБЛИЦЫ METRICS:")
    try:
        from database import SessionLocal
        from sqlalchemy import inspect, text
        
        db = SessionLocal()
        
        # Определяем тип базы данных
        db_url = str(db.bind.url)
        is_postgresql = 'postgresql' in db_url
        
        if is_postgresql:
            print("   📊 Тип БД: PostgreSQL")
        else:
            print("   📊 Тип БД: SQLite")
        
        # Проверяем существующие колонки в таблице metrics
        inspector = inspect(db.bind)
        columns = [col['name'] for col in inspector.get_columns('metrics')]
        print(f"   📋 Существующие колонки: {', '.join(columns)}")
        
        # Добавляем is_global_critical, если нет
        if 'is_global_critical' not in columns:
            print("   ➕ Добавляем колонку is_global_critical...")
            if is_postgresql:
                db.execute(text("ALTER TABLE metrics ADD COLUMN is_global_critical BOOLEAN DEFAULT FALSE"))
            else:
                db.execute(text("ALTER TABLE metrics ADD COLUMN is_global_critical BOOLEAN DEFAULT 0"))
            db.commit()
            print("   ✅ Колонка is_global_critical добавлена")
        else:
            print("   ✅ Колонка is_global_critical уже существует")
        
        # Добавляем allow_na, если нет
        if 'allow_na' not in columns:
            print("   ➕ Добавляем колонку allow_na...")
            if is_postgresql:
                db.execute(text("ALTER TABLE metrics ADD COLUMN allow_na BOOLEAN DEFAULT TRUE"))
            else:
                db.execute(text("ALTER TABLE metrics ADD COLUMN allow_na BOOLEAN DEFAULT 1"))
            db.commit()
            print("   ✅ Колонка allow_na добавлена")
        else:
            print("   ✅ Колонка allow_na уже существует")
        
        db.close()
        print("✅ Миграция таблицы metrics завершена")
        
    except Exception as e:
        print(f"⚠️ Ошибка при миграции metrics: {e}")
        import traceback
        traceback.print_exc()
    # ======================================================
    
    # ============== СОЗДАНИЕ ПОЛЬЗОВАТЕЛЕЙ ==============
    print("\n👤 ПРОВЕРКА ПОЛЬЗОВАТЕЛЕЙ:")
    from database import SessionLocal, User
    from auth import hash_password
    
    db = SessionLocal()
    
    # Создаем админа по умолчанию если нет пользователей
    if db.query(User).count() == 0:
        admin = User(
            full_name="Администратор",
            login="admin",
            password_hash=hash_password("admin123"),
            role="admin",
            avatar_letters="AD",
            is_active=True
        )
        db.add(admin)
        db.commit()
        print("   ✅ Создан пользователь admin / admin123")
    else:
        print("   ✅ Пользователи уже существуют")
    
    # Создаем тестового контролера Яну
    if db.query(User).filter(User.login == "yana.control").count() == 0:
        yana = User(
            full_name="Яна Контролер",
            login="yana.control",
            password_hash=hash_password("yana123"),
            role="controller",
            avatar_letters="ЯК",
            is_active=True
        )
        db.add(yana)
        db.commit()
        print("   ✅ Создан пользователь yana.control / yana123")
    
    # Создаем тестового менеджера
    if db.query(User).filter(User.login == "manager").count() == 0:
        manager = User(
            full_name="Менеджер Проектов",
            login="manager",
            password_hash=hash_password("manager123"),
            role="manager",
            avatar_letters="МП",
            is_active=True
        )
        db.add(manager)
        db.commit()
        print("   ✅ Создан пользователь manager / manager123")
    
    db.close()
    # ======================================================
    
    print("\n" + "="*60)
    print("✅ БАЗА ДАННЫХ ГОТОВА К РАБОТЕ")
    print("="*60 + "\n")

# ============== ГЛАВНЫЙ ДАШБОРД ==============

@app.get("/", response_class=HTMLResponse)
async def dashboard(
    request: Request,
    current_user: Optional[User] = Depends(get_optional_user),
    db: Session = Depends(get_db)
):
    """Главная страница - дашборд с операторами и проектами"""
    if not current_user:
        return RedirectResponse(url="/auth/login", status_code=302)
    
    # Получаем все активные проекты
    projects = db.query(Project).filter(Project.is_active == True).all()
    
    # Получаем всех активных операторов
    operators = db.query(Operator).filter(Operator.is_active == True).all()
    
    # Получаем последние оценки и статистику для операторов
    for operator in operators:
        # Используем среднее качество из поля avg_quality
        operator.last_quality = operator.avg_quality
        
        # Последняя дата оценки
        last_eval = db.query(Evaluation).filter(
            Evaluation.operator_id == operator.id
        ).order_by(Evaluation.call_date.desc()).first()
        operator.last_call_date = last_eval.call_date if last_eval else None
    
    # Статистика по проектам
    projects_stats = []
    for project in projects:
        ops_count = db.query(Operator).filter(
            Operator.project_id == project.id,
            Operator.is_active == True
        ).count()
        
        cutoff_date = date.today() - timedelta(days=30)
        evaluations = db.query(Evaluation).join(Operator).filter(
            Operator.project_id == project.id,
            Evaluation.call_date >= cutoff_date
        ).all()
        
        avg_quality = sum(e.quality_percent for e in evaluations) / len(evaluations) if evaluations else 0
        
        projects_stats.append({
            "id": project.id,
            "name": project.name,
            "operators_count": ops_count,
            "avg_quality": round(avg_quality, 1),
            "evaluations_count": len(evaluations)
        })
    
    # Общая статистика
    total_operators = len(operators)
    total_evaluations = db.query(Evaluation).count()
    today_evaluations = db.query(Evaluation).filter(
        func.date(Evaluation.created_at) == date.today()
    ).count()
    
    # Количество активных тикетов (новые + в работе)
    tickets_count = get_active_tickets_count(db)
    
    # Последние 5 оценок
    recent_evaluations = db.query(Evaluation).order_by(
        Evaluation.created_at.desc()
    ).limit(5).all()
    
    # Получаем контролеров (только для админа и менеджера)
    controllers = []
    if current_user.role in ['admin', 'manager']:
        # Получаем всех пользователей с ролями controller и admin
        controllers = db.query(User).filter(
            User.role.in_(['controller', 'admin']),
            User.is_active == True
        ).all()
        
        # Добавляем статистику для каждого контролера
        for controller in controllers:
            # Количество оценок, сделанных контролером
            controller.evaluations_count = db.query(Evaluation).filter(
                Evaluation.evaluator_id == controller.id
            ).count()
            
            # Дата последней оценки
            last_eval = db.query(Evaluation).filter(
                Evaluation.evaluator_id == controller.id
            ).order_by(Evaluation.created_at.desc()).first()
            
            controller.last_evaluation = last_eval.evaluation_date.strftime('%d.%m.%Y') if last_eval else '—'
            
            # Количество проектов, к которым имеет доступ контролер
            if controller.role == 'admin':
                controller.projects_count = len(projects)
            else:
                controller.projects_count = len(projects)
        
        print(f"👥 Найдено контролеров: {len(controllers)}")
    
    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "user": current_user,
            "projects": projects,
            "projects_stats": projects_stats,
            "operators": operators,
            "total_operators": total_operators,
            "total_evaluations": total_evaluations,
            "today_evaluations": today_evaluations,
            "recent_evaluations": recent_evaluations,
            "tickets_count": tickets_count,
            "controllers": controllers,  
            "date": date
        }
    )

# ============== ОПЕРАТОРЫ ==============

@app.get("/operators", response_class=HTMLResponse)
async def operators_page(
    request: Request,
    project_id: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Страница со списком операторов"""
    query = db.query(Operator).filter(Operator.is_active == True)
    
    if project_id:
        query = query.filter(Operator.project_id == project_id)
    
    operators = query.all()
    projects = db.query(Project).filter(Project.is_active == True).all()
    
    for operator in operators:
        stats = get_operator_stats(operator.id, period_days=30)
        operator.avg_quality_30d = stats['avg_quality'] if stats else 0
        operator.evaluations_count_30d = stats['evaluations_count'] if stats else 0
    
    # Количество активных тикетов (новые + в работе)
    tickets_count = get_active_tickets_count(db)
    
    return templates.TemplateResponse(
        "operators.html",
        {
            "request": request,
            "user": current_user,
            "operators": operators,
            "projects": projects,
            "selected_project": project_id,
            "today": date.today().strftime('%Y-%m-%d'),
            "tickets_count": tickets_count,
            "date": date
        }
    )

@app.get("/operator/{operator_id}", response_class=HTMLResponse)
async def operator_detail(
    request: Request,
    operator_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Детальная страница оператора"""
    operator = db.query(Operator).filter(Operator.id == operator_id).first()
    if not operator:
        raise HTTPException(status_code=404, detail="Оператор не найден")
    
    stats = get_operator_stats(operator_id, period_days=30)
    weekly = get_weekly_dynamics(operator_id=operator_id)
    monthly = get_monthly_dynamics(operator_id=operator_id)
    
    evaluations = db.query(Evaluation).filter(
        Evaluation.operator_id == operator_id
    ).order_by(Evaluation.call_date.desc()).all()
    
    # Получаем метрики через блоки
    metrics = db.query(Metric).join(MetricBlock).filter(
        MetricBlock.project_id == operator.project_id,
        Metric.is_active == True,
        MetricBlock.is_active == True
    ).order_by(MetricBlock.display_order, Metric.display_order).all()
    
    # Получаем все проекты для полосы проектов
    projects = db.query(Project).filter(Project.is_active == True).all()
    
    # Количество активных тикетов (новые + в работе)
    tickets_count = get_active_tickets_count(db)
    
    # Преобразуем в JSON-совместимый формат
    evaluations_json = []
    for eval_obj in evaluations:
        eval_dict = {
            "id": eval_obj.id,
            "call_date": eval_obj.call_date.strftime('%Y-%m-%d'),
            "quality_percent": eval_obj.quality_percent,
            "comment": eval_obj.comment,
            "call_link": eval_obj.call_link,
            "evaluator": {
                "id": eval_obj.evaluator.id,
                "full_name": eval_obj.evaluator.full_name
            },
            "metrics": []
        }
        
        for metric in eval_obj.metrics:
            metric_dict = {
                "id": metric.id,
                "max_score": metric.max_score,
                "earned_score": metric.earned_score,
                "is_not_evaluated": metric.is_not_evaluated,
                "comment": metric.comment,
                "metric": {
                    "id": metric.metric.id,
                    "name": metric.metric.name
                }
            }
            eval_dict["metrics"].append(metric_dict)
        
        evaluations_json.append(eval_dict)
    
    # Преобразуем метрики в словари (без weight)
    metrics_json = []
    for metric in metrics:
        metrics_json.append({
            "id": metric.id,
            "name": metric.name,
            "max_score": metric.max_score,
            "description": metric.description,
            "is_critical": metric.is_critical,
            "resets_block": metric.resets_block,
            "penalty_type": metric.penalty_type,
            "penalty_value": metric.penalty_value
        })
    
    return templates.TemplateResponse(
        "operator_detail.html",
        {
            "request": request,
            "user": current_user,
            "operator": operator,
            "stats": stats,
            "weekly": weekly,
            "monthly": monthly,
            "evaluations": evaluations,
            "evaluations_json": evaluations_json,
            "metrics": metrics,
            "metrics_json": metrics_json,
            "projects": projects,
            "tickets_count": tickets_count,
            "date": date
        }
    )

# ============== ОЦЕНЕННЫЕ ЗВОНКИ ==============

@app.get("/evaluations", response_class=HTMLResponse)
async def evaluations_page(
    request: Request,
    project_id: Optional[str] = None,
    operator_id: Optional[str] = None,
    auditor_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Страница со списком оцененных звонков"""
    
    # Преобразуем строки в числа, если они не пустые
    project_id_int = None
    if project_id and project_id.strip() and project_id.strip().isdigit():
        project_id_int = int(project_id.strip())
    
    operator_id_int = None
    if operator_id and operator_id.strip() and operator_id.strip().isdigit():
        operator_id_int = int(operator_id.strip())
    
    auditor_id_int = None
    if auditor_id and auditor_id.strip() and auditor_id.strip().isdigit():
        auditor_id_int = int(auditor_id.strip())
    
    # Получаем все проекты для фильтра
    projects = db.query(Project).filter(Project.is_active == True).all()
    
    # Получаем всех операторов для фильтра
    operators_query = db.query(Operator).filter(Operator.is_active == True)
    if project_id_int:
        operators_query = operators_query.filter(Operator.project_id == project_id_int)
    operators = operators_query.all()
    
    # Получаем всех аудиторов (пользователей с ролью controller и admin)
    auditors = db.query(User).filter(
        User.role.in_(["controller", "admin"]),
        User.is_active == True
    ).all()
    
    # Базовый запрос для оценок - СОРТИРОВКА ПО НОВИЗНЕ (created_at)
    query = db.query(Evaluation).join(Operator).join(Project)
    
    # Применяем фильтры
    if project_id_int:
        query = query.filter(Operator.project_id == project_id_int)
    
    if operator_id_int:
        query = query.filter(Evaluation.operator_id == operator_id_int)
    
    if auditor_id_int:
        query = query.filter(Evaluation.evaluator_id == auditor_id_int)
    
    if start_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").date()
            query = query.filter(Evaluation.call_date >= start)
        except ValueError:
            pass
    
    if end_date:
        try:
            end = datetime.strptime(end_date, "%Y-%m-%d").date()
            query = query.filter(Evaluation.call_date <= end)
        except ValueError:
            pass
    
    # Получаем оценки с сортировкой по дате создания (новые сверху)
    evaluations = query.order_by(Evaluation.created_at.desc()).all()
    
    # Для каждой оценки получаем связанный тикет
    for evaluation in evaluations:
        evaluation.related_ticket = db.query(Ticket).filter(
            Ticket.evaluation_id == evaluation.id
        ).first()
    
    # Получаем все проекты для полосы проектов
    all_projects = db.query(Project).filter(Project.is_active == True).all()
    
    # Количество активных тикетов для бейджа
    tickets_count = get_active_tickets_count(db)
    
    return templates.TemplateResponse(
        "evaluations.html",
        {
            "request": request,
            "user": current_user,
            "projects": all_projects,
            "evaluations": evaluations,
            "operators": operators,
            "auditors": auditors,
            "selected_project": project_id_int,
            "selected_operator": operator_id_int,
            "selected_auditor": auditor_id_int,
            "start_date": start_date,
            "end_date": end_date,
            "tickets_count": tickets_count,
            "date": date
        }
    )

# ============== УДАЛЕНИЕ ОЦЕНКИ ==============

@app.delete("/api/evaluations/{evaluation_id}")
async def delete_evaluation(
    evaluation_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Удаление оценки звонка"""
    try:
        # Получаем оценку
        evaluation = db.query(Evaluation).filter(Evaluation.id == evaluation_id).first()
        if not evaluation:
            return JSONResponse(
                status_code=404,
                content={"status": "error", "message": "Оценка не найдена"}
            )
        
        # Проверяем права (админ или автор оценки)
        if current_user.role != 'admin' and current_user.id != evaluation.evaluator_id:
            return JSONResponse(
                status_code=403,
                content={"status": "error", "message": "Нет прав на удаление этой оценки"}
            )
        
        # Сохраняем ID оператора для обновления статистики
        operator_id = evaluation.operator_id
        
        # Удаляем связанные записи (каскадно должны удалиться, но для надежности)
        # Сначала удаляем метрики оценки
        db.query(EvaluationMetric).filter(EvaluationMetric.evaluation_id == evaluation_id).delete()
        
        # Удаляем результаты блоков
        db.query(EvaluationBlockResult).filter(EvaluationBlockResult.evaluation_id == evaluation_id).delete()
        
        # Удаляем саму оценку
        db.delete(evaluation)
        
        # Обновляем статистику оператора
        update_operator_avg_quality(operator_id, db)
        
        db.commit()
        
        return JSONResponse({
            "status": "success",
            "message": "Оценка успешно удалена"
        })
        
    except Exception as e:
        db.rollback()
        print(f"❌ Ошибка при удалении оценки: {str(e)}")
        return JSONResponse(
            status_code=500,
            content={"status": "error", "message": f"Внутренняя ошибка сервера: {str(e)}"}
        )

# ============== МАССОВОЕ УДАЛЕНИЕ ОЦЕНОК ==============

@app.post("/api/evaluations/bulk-delete")
async def bulk_delete_evaluations(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Массовое удаление оценок"""
    try:
        data = await request.json()
        evaluation_ids = data.get("evaluation_ids", [])
        
        if not evaluation_ids:
            return JSONResponse(
                status_code=400,
                content={"status": "error", "message": "Не указаны ID оценок"}
            )
        
        # Проверяем права для каждой оценки
        deleted_count = 0
        failed_count = 0
        operator_ids = set()
        
        for eval_id in evaluation_ids:
            evaluation = db.query(Evaluation).filter(Evaluation.id == eval_id).first()
            if not evaluation:
                failed_count += 1
                continue
            
            # Проверяем права (админ или автор оценки)
            if current_user.role != 'admin' and current_user.id != evaluation.evaluator_id:
                failed_count += 1
                continue
            
            operator_ids.add(evaluation.operator_id)
            
            # Удаляем связанные записи
            db.query(EvaluationMetric).filter(EvaluationMetric.evaluation_id == eval_id).delete()
            db.query(EvaluationBlockResult).filter(EvaluationBlockResult.evaluation_id == eval_id).delete()
            db.delete(evaluation)
            deleted_count += 1
        
        # Обновляем статистику для всех операторов
        for op_id in operator_ids:
            update_operator_avg_quality(op_id, db)
        
        db.commit()
        
        return JSONResponse({
            "status": "success",
            "message": f"Удалено оценок: {deleted_count}, не удалось: {failed_count}",
            "deleted": deleted_count,
            "failed": failed_count
        })
        
    except Exception as e:
        db.rollback()
        print(f"❌ Ошибка при массовом удалении: {str(e)}")
        return JSONResponse(
            status_code=500,
            content={"status": "error", "message": f"Внутренняя ошибка сервера: {str(e)}"}
        )

# ============== ЭКСПОРТ В EXCEL ==============

# ============== ЭКСПОРТ В EXCEL ==============

@app.get("/api/evaluations/export")
async def export_evaluations_to_excel(
    request: Request,
    project_id: Optional[str] = None,
    operator_id: Optional[str] = None,
    auditor_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Экспорт оцененных звонков в Excel с детальной информацией"""
    
    if not EXCEL_AVAILABLE:
        raise HTTPException(
            status_code=500, 
            detail="Библиотеки для Excel не установлены. Установите: pip install pandas openpyxl"
        )
    
    try:
        # Преобразуем строки в числа, если они не пустые
        project_id_int = None
        if project_id and project_id.strip() and project_id.strip().isdigit():
            project_id_int = int(project_id.strip())
        
        operator_id_int = None
        if operator_id and operator_id.strip() and operator_id.strip().isdigit():
            operator_id_int = int(operator_id.strip())
        
        auditor_id_int = None
        if auditor_id and auditor_id.strip() and auditor_id.strip().isdigit():
            auditor_id_int = int(auditor_id.strip())
        
        # Базовый запрос для оценок
        query = db.query(Evaluation).join(Operator).join(Project).join(User, Evaluation.evaluator_id == User.id)
        
        # Применяем фильтры
        if project_id_int:
            query = query.filter(Operator.project_id == project_id_int)
        
        if operator_id_int:
            query = query.filter(Evaluation.operator_id == operator_id_int)
        
        if auditor_id_int:
            query = query.filter(Evaluation.evaluator_id == auditor_id_int)
        
        if start_date:
            try:
                start = datetime.strptime(start_date, "%Y-%m-%d").date()
                query = query.filter(Evaluation.call_date >= start)
            except ValueError:
                pass
        
        if end_date:
            try:
                end = datetime.strptime(end_date, "%Y-%m-%d").date()
                query = query.filter(Evaluation.call_date <= end)
            except ValueError:
                pass
        
        # Получаем оценки с сортировкой
        evaluations = query.order_by(Evaluation.call_date.desc()).all()
        
        if not evaluations:
            # Если нет данных, возвращаем пустой Excel с сообщением
            wb = Workbook()
            ws = wb.active
            ws.title = "Оценки"
            ws.append(["Нет данных за выбранный период"])
            
            filename = f"evaluations_empty_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return Response(
                content=save_virtual_workbook(wb),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        
        # Создаем новую книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Оценки звонков"
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Получаем все уникальные блоки и метрики для создания колонок
        all_blocks = {}
        all_metrics = {}
        
        for evaluation in evaluations:
            for block_result in evaluation.block_results:
                if block_result.block_id not in all_blocks:
                    all_blocks[block_result.block_id] = block_result.block.name
            
            for metric in evaluation.metrics:
                if metric.metric_id not in all_metrics and metric.metric:
                    all_metrics[metric.metric_id] = {
                        'name': metric.metric.name,
                        'block_id': metric.metric.block_id,
                        'block_name': metric.metric.block.name if metric.metric.block else 'Без блока'
                    }
        
        # Сортируем блоки и метрики
        sorted_blocks = sorted(all_blocks.items(), key=lambda x: x[1])
        sorted_metrics = sorted(all_metrics.items(), key=lambda x: (x[1]['block_name'], x[1]['name']))
        
        # Формируем заголовки
        headers = [
            "ID оценки",
            "Дата оценки",
            "Дата звонка",
            "Аудитор",
            "Оператор",
            "Проект",
            "Тема звонка",
            "Длительность",
            "Ссылка на запись",
            "Общий комментарий",
            "Общий % качества",
            "Набранные баллы",
            "Макс. баллы"
        ]
        
        # Добавляем заголовки для каждого блока (проценты)
        block_columns = {}
        for block_id, block_name in sorted_blocks:
            col_name = f"{block_name} (%)"
            headers.append(col_name)
            block_columns[block_id] = col_name
        
        # Добавляем заголовки для опциональности блоков
        optional_columns = {}
        for block_id, block_name in sorted_blocks:
            col_name = f"{block_name} (обязателен)"
            headers.append(col_name)
            optional_columns[block_id] = col_name
        
        # Добавляем заголовки для каждой метрики (ДА/НЕТ/Н/О)
        metric_columns = {}
        for metric_id, metric_info in sorted_metrics:
            col_name = f"{metric_info['block_name']}: {metric_info['name']}"
            headers.append(col_name)
            metric_columns[metric_id] = col_name
        
        # Добавляем заголовки для комментариев к метрикам
        comment_columns = {}
        for metric_id, metric_info in sorted_metrics:
            col_name = f"Комментарий: {metric_info['name']}"
            headers.append(col_name)
            comment_columns[metric_id] = col_name
        
        # Добавляем заголовки для числителя и знаменателя по каждому блоку
        numerator_columns = {}
        denominator_columns = {}
        for block_id, block_name in sorted_blocks:
            numerator_col = f"{block_name} (баллы)"
            denominator_col = f"{block_name} (макс)"
            headers.append(numerator_col)
            headers.append(denominator_col)
            numerator_columns[block_id] = numerator_col
            denominator_columns[block_id] = denominator_col
        
        # Добавляем информацию о тикете
        headers.extend([
            "Тикет создан",
            "Статус тикета",
            "Ссылка на тикет"
        ])
        
        # Записываем заголовки
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        # Записываем данные
        for row_num, evaluation in enumerate(evaluations, 2):
            col_num = 1
            
            # Получаем связанный тикет
            related_ticket = db.query(Ticket).filter(Ticket.evaluation_id == evaluation.id).first()
            
            # Основная информация
            ws.cell(row=row_num, column=col_num, value=evaluation.id).border = border
            col_num += 1
            
            ws.cell(row=row_num, column=col_num, value=evaluation.evaluation_date.strftime('%d.%m.%Y') if evaluation.evaluation_date else '').border = border
            col_num += 1
            
            ws.cell(row=row_num, column=col_num, value=evaluation.call_date.strftime('%d.%m.%Y') if evaluation.call_date else '').border = border
            col_num += 1
            
            ws.cell(row=row_num, column=col_num, value=evaluation.evaluator.full_name if evaluation.evaluator else '').border = border
            col_num += 1
            
            ws.cell(row=row_num, column=col_num, value=evaluation.operator.full_name if evaluation.operator else '').border = border
            col_num += 1
            
            ws.cell(row=row_num, column=col_num, value=evaluation.operator.project.name if evaluation.operator and evaluation.operator.project else '').border = border
            col_num += 1
            
            ws.cell(row=row_num, column=col_num, value=evaluation.topic.name if evaluation.topic else '').border = border
            col_num += 1
            
            # Длительность звонка
            if evaluation.call_duration:
                minutes = evaluation.call_duration // 60
                seconds = evaluation.call_duration % 60
                ws.cell(row=row_num, column=col_num, value=f"{minutes}:{seconds:02d}").border = border
            else:
                ws.cell(row=row_num, column=col_num, value="—").border = border
            col_num += 1
            
            ws.cell(row=row_num, column=col_num, value=evaluation.call_link or '').border = border
            col_num += 1
            
            ws.cell(row=row_num, column=col_num, value=evaluation.comment or '').border = border
            col_num += 1
            
            ws.cell(row=row_num, column=col_num, value=round(evaluation.quality_percent or 0, 1)).border = border
            col_num += 1
            
            # ОБЩИЕ БАЛЛЫ - ИСКЛЮЧАЕМ Н/О МЕТРИКИ
            total_earned = 0
            total_max = 0
            for metric in evaluation.metrics:
                if not metric.is_not_evaluated:  # Исключаем Н/О метрики
                    total_earned += metric.earned_score or 0
                    total_max += metric.max_score or 0
            
            ws.cell(row=row_num, column=col_num, value=total_earned).border = border
            col_num += 1
            ws.cell(row=row_num, column=col_num, value=total_max).border = border
            col_num += 1
            
            # Проценты по блокам
            block_results_dict = {br.block_id: br for br in evaluation.block_results}
            for block_id, _ in sorted_blocks:
                block_result = block_results_dict.get(block_id)
                percent = round(block_result.percent, 1) if block_result else 0
                ws.cell(row=row_num, column=col_num, value=percent).border = border
                col_num += 1
            
            # Опциональность блоков (ДА/НЕТ) - заглушка, нужно будет добавить в модель
            # Пока ставим "НЕТ" как обязательные
            for block_id, _ in sorted_blocks:
                ws.cell(row=row_num, column=col_num, value="НЕТ").border = border
                col_num += 1
            
            # Значения метрик (ДА/НЕТ/Н/О)
            metrics_dict = {m.metric_id: m for m in evaluation.metrics}
            for metric_id, metric_info in sorted_metrics:
                metric = metrics_dict.get(metric_id)
                if metric:
                    if metric.is_not_evaluated:
                        ws.cell(row=row_num, column=col_num, value="Н/О").border = border
                    elif metric.earned_score and metric.earned_score == metric.max_score:
                        ws.cell(row=row_num, column=col_num, value="ДА").border = border
                    elif metric.earned_score == 0:
                        ws.cell(row=row_num, column=col_num, value="НЕТ").border = border
                    else:
                        ws.cell(row=row_num, column=col_num, value=metric.earned_score).border = border
                else:
                    ws.cell(row=row_num, column=col_num, value="—").border = border
                col_num += 1
            
            # Комментарии к метрикам
            for metric_id, metric_info in sorted_metrics:
                metric = metrics_dict.get(metric_id)
                ws.cell(row=row_num, column=col_num, value=metric.comment if metric and metric.comment else '').border = border
                col_num += 1
            
            # Числитель и знаменатель по блокам
            for block_id, _ in sorted_blocks:
                block_result = block_results_dict.get(block_id)
                earned = round(block_result.earned_score, 1) if block_result else 0
                max_score = round(block_result.max_score, 1) if block_result else 0
                ws.cell(row=row_num, column=col_num, value=earned).border = border
                col_num += 1
                ws.cell(row=row_num, column=col_num, value=max_score).border = border
                col_num += 1
            
            # Информация о тикете
            if related_ticket:
                ws.cell(row=row_num, column=col_num, value="ДА").border = border
                col_num += 1
                
                status_text = {
                    'new': '🆕 Новый',
                    'in_progress': '🔄 В работе',
                    'completed': '✅ Завершен',
                    'cancelled': '❌ Отменен'
                }.get(related_ticket.status, related_ticket.status)
                ws.cell(row=row_num, column=col_num, value=status_text).border = border
                col_num += 1
                
                ws.cell(row=row_num, column=col_num, value=f"/tickets/{related_ticket.id}").border = border
                col_num += 1
            else:
                ws.cell(row=row_num, column=col_num, value="НЕТ").border = border
                col_num += 1
                ws.cell(row=row_num, column=col_num, value="—").border = border
                col_num += 1
                ws.cell(row=row_num, column=col_num, value="—").border = border
                col_num += 1
        
        # Авто-подбор ширины колонок
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, len(evaluations) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем в буфер
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Формируем имя файла
        filename = f"evaluations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Возвращаем файл
        return Response(
            content=excel_file.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        print(f"❌ Ошибка при экспорте в Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Ошибка при экспорте: {str(e)}")

# ============== API ДЛЯ ПОЛУЧЕНИЯ АУДИТОРОВ ==============

@app.get("/api/auditors")
async def get_auditors(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получение списка аудиторов (контролеров и админов)"""
    auditors = db.query(User).filter(
        User.role.in_(["controller", "admin"]),
        User.is_active == True
    ).order_by(User.full_name).all()
    
    result = []
    for auditor in auditors:
        # Получаем количество оценок
        evaluations_count = db.query(Evaluation).filter(
            Evaluation.evaluator_id == auditor.id
        ).count()
        
        result.append({
            "id": auditor.id,
            "full_name": auditor.full_name,
            "login": auditor.login,
            "role": auditor.role,
            "avatar_letters": auditor.avatar_letters,
            "evaluations_count": evaluations_count
        })
    
    return JSONResponse(result)

# ============== СТАТИСТИКА ПО КОНТРОЛЕРАМ ==============

@app.get("/controllers", response_class=HTMLResponse)
async def controllers_stats_page(
    request: Request,
    controller_id: Optional[str] = Query(None, description="Фильтр по контролеру"),
    project_id: Optional[str] = Query(None, description="Фильтр по проекту"),
    start_date: Optional[str] = Query(None, description="Начальная дата"),
    end_date: Optional[str] = Query(None, description="Конечная дата"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Страница со статистикой по контролерам"""
    
    # Проверка прав (только админ и руководитель)
    if current_user.role not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    
    # Преобразуем ID в числа, если они не пустые
    controller_id_int = None
    if controller_id and controller_id.strip():
        try:
            controller_id_int = int(controller_id)
        except ValueError:
            controller_id_int = None
    
    project_id_int = None
    if project_id and project_id.strip():
        try:
            project_id_int = int(project_id)
        except ValueError:
            project_id_int = None
    
    # Получаем всех контролеров (включая админов как контролеров)
    controllers = db.query(User).filter(
        User.role.in_(['controller', 'admin']),
        User.is_active == True
    ).order_by(User.full_name).all()
    
    # Получаем все проекты для фильтра
    projects = db.query(Project).filter(Project.is_active == True).all()
    
    # Получаем все блоки проектов
    blocks_query = db.query(MetricBlock).filter(MetricBlock.is_active == True)
    if project_id_int:
        blocks_query = blocks_query.filter(MetricBlock.project_id == project_id_int)
    blocks = blocks_query.order_by(MetricBlock.display_order).all()
    
    # Определяем диапазон дат
    date_start = None
    date_end = None
    date_range_days = 0
    all_dates = []
    
    if start_date:
        try:
            date_start = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            pass
    else:
        # По умолчанию последние 30 дней
        date_end = date.today()
        date_start = date_end - timedelta(days=29)
        start_date = date_start.strftime("%Y-%m-%d")
        end_date = date_end.strftime("%Y-%m-%d")
    
    if end_date:
        try:
            date_end = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            pass
    
    if date_start and date_end:
        date_range_days = (date_end - date_start).days + 1
        current = date_start
        while current <= date_end:
            all_dates.append({
                'date': current.strftime('%Y-%m-%d'),
                'label': current.strftime('%d.%m')
            })
            current += timedelta(days=1)
    
    # Базовый запрос для оценок
    query = db.query(
        Evaluation,
        User.full_name.label('controller_name'),
        Project.name.label('project_name'),
        func.date(Evaluation.created_at).label('eval_date')
    ).join(User, Evaluation.evaluator_id == User.id).join(Operator).join(Project)
    
    # Применяем фильтры
    if controller_id_int:
        query = query.filter(Evaluation.evaluator_id == controller_id_int)
    
    if project_id_int:
        query = query.filter(Operator.project_id == project_id_int)
    
    if date_start:
        query = query.filter(func.date(Evaluation.created_at) >= date_start)
    
    if date_end:
        query = query.filter(func.date(Evaluation.created_at) <= date_end)
    
    # Получаем все оценки
    evaluations = query.order_by(Evaluation.created_at.desc()).all()
    
    # Группируем по дням для графиков
    daily_stats = {}
    daily_quality = {}
    daily_blocks = {}
    
    # Инициализируем все даты нулями
    for date_info in all_dates:
        date_str = date_info['date']
        daily_stats[date_str] = {
            'date': date_str,
            'label': date_info['label'],
            'count': 0
        }
        daily_quality[date_str] = {
            'date': date_str,
            'label': date_info['label'],
            'total_quality': 0,
            'count': 0
        }
    
    # Собираем данные по блокам
    for block in blocks:
        daily_blocks[block.id] = {
            'block_id': block.id,
            'block_name': block.name,
            'data': {}
        }
        for date_info in all_dates:
            date_str = date_info['date']
            daily_blocks[block.id]['data'][date_str] = {
                'date': date_str,
                'label': date_info['label'],
                'total_percent': 0,
                'count': 0
            }
    
    # Обрабатываем оценки
    for eval_obj, controller_name, project_name, eval_date in evaluations:
        # Преобразуем eval_date в объект date, если это строка
        if isinstance(eval_date, str):
            try:
                eval_date_obj = datetime.strptime(eval_date, "%Y-%m-%d").date()
            except ValueError:
                continue
        else:
            eval_date_obj = eval_date
            
        date_str = eval_date_obj.strftime('%Y-%m-%d')
        
        # Обновляем статистику по дням
        if date_str in daily_stats:
            daily_stats[date_str]['count'] += 1
        
        if date_str in daily_quality:
            daily_quality[date_str]['total_quality'] += eval_obj.quality_percent
            daily_quality[date_str]['count'] += 1
        
        # Обновляем статистику по блокам
        for br in eval_obj.block_results:
            block_id = br.block_id
            if block_id in daily_blocks and date_str in daily_blocks[block_id]['data']:
                daily_blocks[block_id]['data'][date_str]['total_percent'] += br.percent
                daily_blocks[block_id]['data'][date_str]['count'] += 1
    
    # Преобразуем в списки для графиков
    count_chart_data = []
    quality_chart_data = []
    blocks_chart_data = []
    
    # Сортируем даты
    sorted_dates = sorted(all_dates, key=lambda x: x['date'])
    
    for date_info in sorted_dates:
        date_str = date_info['date']
        
        # Данные для графика количества
        count_chart_data.append({
            'label': date_info['label'],
            'count': daily_stats[date_str]['count']
        })
        
        # Данные для графика качества
        qual_stats = daily_quality[date_str]
        avg_quality = 0
        if qual_stats['count'] > 0:
            avg_quality = qual_stats['total_quality'] / qual_stats['count']
        quality_chart_data.append({
            'label': date_info['label'],
            'quality': round(avg_quality, 1)
        })
    
    # Данные для графиков по блокам
    for block_id, block_data in daily_blocks.items():
        block_chart_data = []
        for date_info in sorted_dates:
            date_str = date_info['date']
            day_data = block_data['data'][date_str]
            avg_percent = 0
            if day_data['count'] > 0:
                avg_percent = day_data['total_percent'] / day_data['count']
            block_chart_data.append({
                'label': date_info['label'],
                'percent': round(avg_percent, 1)
            })
        
        blocks_chart_data.append({
            'block_id': block_id,
            'block_name': block_data['block_name'],
            'data': block_chart_data
        })
    
    # Данные для таблицы экспорта
    table_data = []
    for eval_obj, controller_name, project_name, eval_date in evaluations:
        # Преобразуем eval_date в объект date, если это строка
        if isinstance(eval_date, str):
            try:
                eval_date_obj = datetime.strptime(eval_date, "%Y-%m-%d").date()
                date_formatted = eval_date_obj.strftime('%d.%m.%Y')
            except ValueError:
                date_formatted = eval_date
        else:
            date_formatted = eval_date.strftime('%d.%m.%Y')
        
        table_data.append({
            'date': date_formatted,
            'project': project_name,
            'controller': controller_name,
            'operator': eval_obj.operator.full_name,
            'quality': round(eval_obj.quality_percent, 1)
        })
    
    # Общая статистика
    total_evaluations = len(evaluations)
    unique_controllers = len(set(e[1] for e in evaluations)) if evaluations else 0
    
    # Количество активных тикетов для бейджа
    tickets_count = get_active_tickets_count(db)
    
    return templates.TemplateResponse(
        "controllers_stats.html",
        {
            "request": request,
            "user": current_user,
            "controllers": controllers,
            "projects": projects,
            "selected_controller": controller_id_int,
            "selected_project": project_id_int,
            "start_date": start_date,
            "end_date": end_date,
            "count_chart_data": count_chart_data,
            "quality_chart_data": quality_chart_data,
            "blocks_chart_data": blocks_chart_data,
            "all_dates": all_dates,
            "date_range_days": date_range_days,
            "table_data": table_data,
            "total_evaluations": total_evaluations,
            "unique_controllers": unique_controllers,
            "tickets_count": tickets_count,
            "date": date
        }
    )

# ============== ЭКСПОРТ СТАТИСТИКИ ПО КОНТРОЛЕРАМ ==============

@app.get("/api/controllers/export")
async def export_controllers_stats(
    request: Request,
    controller_id: Optional[str] = Query(None, description="Фильтр по контролеру"),
    project_id: Optional[str] = Query(None, description="Фильтр по проекту"),
    start_date: Optional[str] = Query(None, description="Начальная дата"),
    end_date: Optional[str] = Query(None, description="Конечная дата"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Экспорт статистики по контролерам в Excel"""
    
    if not EXCEL_AVAILABLE:
        raise HTTPException(
            status_code=500, 
            detail="Библиотеки для Excel не установлены. Установите: pip install pandas openpyxl"
        )
    
    # Проверка прав
    if current_user.role not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    
    try:
        # Преобразуем ID в числа, если они не пустые
        controller_id_int = None
        if controller_id and controller_id.strip():
            try:
                controller_id_int = int(controller_id)
            except ValueError:
                controller_id_int = None
        
        project_id_int = None
        if project_id and project_id.strip():
            try:
                project_id_int = int(project_id)
            except ValueError:
                project_id_int = None
        
        # Базовый запрос для оценок
        query = db.query(
            Evaluation,
            User.full_name.label('controller_name'),
            Project.name.label('project_name'),
            func.date(Evaluation.created_at).label('eval_date')
        ).join(User, Evaluation.evaluator_id == User.id).join(Operator).join(Project)
        
        # Применяем фильтры
        if controller_id_int:
            query = query.filter(Evaluation.evaluator_id == controller_id_int)
        
        if project_id_int:
            query = query.filter(Operator.project_id == project_id_int)
        
        if start_date:
            try:
                start = datetime.strptime(start_date, "%Y-%m-%d").date()
                query = query.filter(func.date(Evaluation.created_at) >= start)
            except ValueError:
                pass
        
        if end_date:
            try:
                end = datetime.strptime(end_date, "%Y-%m-%d").date()
                query = query.filter(func.date(Evaluation.created_at) <= end)
            except ValueError:
                pass
        
        # Получаем все оценки
        evaluations = query.order_by(Evaluation.created_at.desc()).all()
        
        # Группируем по дням для дневной статистики
        daily_data = {}
        for eval_obj, controller_name, project_name, eval_date in evaluations:
            # Преобразуем eval_date в объект date, если это строка
            if isinstance(eval_date, str):
                try:
                    eval_date_obj = datetime.strptime(eval_date, "%Y-%m-%d").date()
                    date_str = eval_date_obj.strftime('%d.%m.%Y')
                except ValueError:
                    date_str = eval_date
            else:
                date_str = eval_date.strftime('%d.%m.%Y')
            
            key = (date_str, project_name, controller_name)
            
            if key not in daily_data:
                daily_data[key] = {
                    'date': date_str,
                    'project': project_name,
                    'controller': controller_name,
                    'count': 0,
                    'total_quality': 0
                }
            
            daily_data[key]['count'] += 1
            daily_data[key]['total_quality'] += eval_obj.quality_percent
        
        # Формируем данные для экспорта
        export_data = []
        for (date_str, project_name, controller_name), stats in daily_data.items():
            avg_quality = stats['total_quality'] / stats['count'] if stats['count'] > 0 else 0
            export_data.append({
                'Дата': date_str,
                'Проект': project_name,
                'Контролер': controller_name,
                'Количество оценок': stats['count'],
                'Средняя оценка': f"{avg_quality:.1f}%"
            })
        
        # Сортируем по дате
        export_data.sort(key=lambda x: x['Дата'])
        
        # Создаем DataFrame
        df = pd.DataFrame(export_data)
        
        # Добавляем итоговую строку
        total_evaluations = sum(row['Количество оценок'] for row in export_data)
        total_row = pd.DataFrame({
            'Дата': ['ИТОГО:'],
            'Проект': [''],
            'Контролер': [''],
            'Количество оценок': [total_evaluations],
            'Средняя оценка': ['']
        })
        df = pd.concat([df, total_row], ignore_index=True)
        
        # Создаем Excel файл
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Статистика по контролерам', index=False)
            
            # Настраиваем ширину колонок
            worksheet = writer.sheets['Статистика по контролерам']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # Генерируем имя файла
        filename = f"controllers_stats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except Exception as e:
        print(f"❌ Ошибка при экспорте статистики контролеров: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Ошибка при экспорте: {str(e)}")

# ============== API ДЛЯ РЕДАКТИРОВАНИЯ ДЛИТЕЛЬНОСТИ ==============

@app.post("/api/evaluations/{evaluation_id}/duration")
async def update_evaluation_duration(
    evaluation_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Обновление длительности звонка"""
    # Получаем оценку
    evaluation = db.query(Evaluation).filter(Evaluation.id == evaluation_id).first()
    if not evaluation:
        raise HTTPException(status_code=404, detail="Оценка не найдена")
    
    # Проверяем права (админ или автор оценки)
    if current_user.role != 'admin' and current_user.id != evaluation.evaluator_id:
        raise HTTPException(status_code=403, detail="Нет прав на редактирование этой оценки")
    
    # Получаем данные из запроса
    data = await request.json()
    duration = data.get("duration")
    
    if duration is None:
        raise HTTPException(status_code=400, detail="Не указана длительность")
    
    # Обновляем длительность (в секундах)
    evaluation.call_duration = duration
    
    # Обновляем среднюю длительность оператора
    update_operator_avg_quality(evaluation.operator_id, db)
    
    db.commit()
    
    return JSONResponse({
        "status": "success",
        "message": "Длительность обновлена",
        "duration": duration
    })

# ============== РЕДАКТИРОВАНИЕ ОЦЕНКИ ==============

@app.get("/evaluate/{operator_id}/edit", response_class=HTMLResponse)
async def edit_evaluation_page(
    request: Request,
    operator_id: int,
    evaluation_id: int = Query(..., description="ID оценки для редактирования"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Страница редактирования оценки звонка"""
    if current_user.role not in ['admin', 'controller']:
        raise HTTPException(status_code=403, detail="Только контролеры могут редактировать оценки")
    
    # Получаем оператора
    operator = db.query(Operator).filter(Operator.id == operator_id).first()
    if not operator:
        raise HTTPException(status_code=404, detail="Оператор не найден")
    
    # Получаем оценку для редактирования
    evaluation = db.query(Evaluation).filter(
        Evaluation.id == evaluation_id,
        Evaluation.operator_id == operator_id
    ).first()
    
    if not evaluation:
        raise HTTPException(status_code=404, detail="Оценка не найдена")
    
    # Проверяем права на редактирование
    if current_user.role != 'admin' and evaluation.evaluator_id != current_user.id:
        raise HTTPException(status_code=403, detail="Нельзя редактировать чужую оценку")
    
    # Получаем метрики через блоки проекта
    metrics = db.query(Metric).join(MetricBlock).filter(
        MetricBlock.project_id == operator.project_id,
        Metric.is_active == True,
        MetricBlock.is_active == True
    ).order_by(MetricBlock.display_order, Metric.display_order).all()
    
    # Получаем блоки для группировки
    blocks = db.query(MetricBlock).filter(
        MetricBlock.project_id == operator.project_id,
        MetricBlock.is_active == True
    ).order_by(MetricBlock.display_order).all()
    
    # Получаем темы проекта
    topics = db.query(Topic).filter(
        Topic.project_id == operator.project_id,
        Topic.is_active == True
    ).order_by(Topic.display_order).all()
    
    # Получаем все проекты для полосы проектов
    projects = db.query(Project).filter(Project.is_active == True).all()
    
    # Количество активных тикетов
    tickets_count = get_active_tickets_count(db)
    
    # Собираем значения метрик для предзаполнения
    metric_values = {}
    for em in evaluation.metrics:
        metric_values[em.metric_id] = {
            "earned_score": em.earned_score,
            "comment": em.comment,
            "is_not_evaluated": em.is_not_evaluated
        }
    
    return templates.TemplateResponse(
        "edit_evaluation.html",
        {
            "request": request,
            "user": current_user,
            "operator": operator,
            "evaluation": evaluation,
            "metrics": metrics,
            "blocks": blocks,
            "topics": topics,
            "metric_values": metric_values,
            "projects": projects,
            "tickets_count": tickets_count,
            "today": date.today().strftime("%Y-%m-%d"),
            "date": date
        }
    )

@app.post("/api/evaluation/update")
async def update_evaluation(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Обновление существующей оценки"""
    # Проверка прав
    if current_user.role not in ['admin', 'controller']:
        raise HTTPException(status_code=403, detail="Только контролеры могут редактировать оценки")
    
    # Получаем данные
    try:
        data = await request.json()
        print(f"\n{'='*60}")
        print(f"📝 ПОЛУЧЕН ЗАПРОС НА ОБНОВЛЕНИЕ ОЦЕНКИ")
        print(f"{'='*60}")
        print(json.dumps(data, indent=2, ensure_ascii=False))
    except Exception as e:
        print(f"❌ Ошибка парсинга JSON: {e}")
        raise HTTPException(status_code=400, detail="Неверный формат данных")
    
    # Проверяем обязательные поля
    evaluation_id = data.get("evaluation_id")
    if not evaluation_id:
        raise HTTPException(status_code=400, detail="evaluation_id обязателен")
    
    operator_id = data.get("operator_id")
    if not operator_id:
        raise HTTPException(status_code=400, detail="operator_id обязателен")
    
    # Получаем оценку из БД
    evaluation = db.query(Evaluation).filter(Evaluation.id == evaluation_id).first()
    if not evaluation:
        raise HTTPException(status_code=404, detail="Оценка не найдена")
    
    # Проверяем права на редактирование
    if current_user.role != 'admin' and evaluation.evaluator_id != current_user.id:
        raise HTTPException(status_code=403, detail="Нельзя редактировать чужую оценку")
    
    # Получаем даты
    call_date = data.get("call_date")
    if not call_date:
        raise HTTPException(status_code=400, detail="call_date обязателен")
    
    try:
        call_date_obj = datetime.strptime(call_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Неверный формат даты звонка")
    
    evaluation_date = data.get("evaluation_date")
    try:
        evaluation_date_obj = datetime.strptime(evaluation_date, "%Y-%m-%d").date() if evaluation_date else date.today()
    except ValueError:
        evaluation_date_obj = date.today()
    
    call_link = data.get("call_link", "")
    comment = data.get("comment", "")
    metrics_data = data.get("metrics", [])
    topic_id = data.get("topic_id")
    
    # Получаем оператора
    operator = db.query(Operator).filter(Operator.id == operator_id).first()
    if not operator:
        raise HTTPException(status_code=404, detail="Оператор не найден")
    
    try:
        # Обновляем основные поля оценки
        evaluation.call_date = call_date_obj
        evaluation.evaluation_date = evaluation_date_obj
        evaluation.call_link = call_link
        evaluation.comment = comment
        evaluation.topic_id = topic_id
        
        # Удаляем старые метрики
        for em in evaluation.metrics:
            db.delete(em)
        db.commit()
        
        # Добавляем новые метрики
        for m in metrics_data:
            metric = db.query(Metric).filter(Metric.id == m["metric_id"]).first()
            if not metric:
                continue
            
            metric_comment = m.get("comment", "")
            earned_score = m.get("earned_score")
            
            eval_metric = EvaluationMetric(
                evaluation_id=evaluation.id,
                metric_id=m["metric_id"],
                earned_score=earned_score,
                max_score=metric.max_score,
                comment=metric_comment,
                is_not_evaluated=m.get("is_not_evaluated", False)
            )
            db.add(eval_metric)
        
        db.flush()
        
        # Пересчитываем проценты
        new_quality = calculate_evaluation_scores(evaluation, db)
        
        # Обновляем статистику оператора
        all_evaluations = db.query(Evaluation).filter(
            Evaluation.operator_id == operator.id
        ).all()
        
        if all_evaluations:
            operator.avg_quality = sum(e.quality_percent for e in all_evaluations) / len(all_evaluations)
            operator.total_evaluations = len(all_evaluations)
        
        # Сохраняем изменения
        db.commit()
        
        return JSONResponse({
            "status": "success",
            "evaluation_id": evaluation.id,
            "quality_percent": round(evaluation.quality_percent, 1),
            "message": "Оценка успешно обновлена"
        })
        
    except Exception as e:
        db.rollback()
        print(f"\n❌ ОШИБКА ОБНОВЛЕНИЯ ОЦЕНКИ: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Ошибка обновления: {str(e)}")

# ============== API ДЛЯ СТАТИСТИКИ ПРОЕКТА ==============

@app.get("/api/project/{project_id}/stats")
async def get_project_stats_api(
    project_id: int,
    days: int = Query(30, description="Количество дней для анализа"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получение расширенной статистики по проекту"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Проект не найден")
    
    # Получаем операторов проекта
    operators = db.query(Operator).filter(
        Operator.project_id == project_id,
        Operator.is_active == True
    ).all()
    
    # Получаем оценки за период
    cutoff_date = date.today() - timedelta(days=days)
    evaluations = db.query(Evaluation).join(Operator).filter(
        Operator.project_id == project_id,
        Evaluation.call_date >= cutoff_date
    ).all()
    
    # Статистика по блокам
    blocks_stats = []
    blocks = db.query(MetricBlock).filter(
        MetricBlock.project_id == project_id,
        MetricBlock.is_active == True
    ).all()
    
    for block in blocks:
        block_results = db.query(EvaluationBlockResult).join(
            Evaluation, Evaluation.id == EvaluationBlockResult.evaluation_id
        ).filter(
            EvaluationBlockResult.block_id == block.id,
            Evaluation.call_date >= cutoff_date
        ).all()
        
        if block_results:
            avg_percent = sum(r.percent for r in block_results) / len(block_results)
            resets_count = sum(1 for r in block_results if r.was_reset)
        else:
            avg_percent = 0
            resets_count = 0
        
        blocks_stats.append({
            "id": block.id,
            "name": block.name,
            "avg_percent": round(avg_percent, 1),
            "evaluations_count": len(block_results),
            "resets_count": resets_count
        })
    
    # Топ операторов
    top_operators = []
    for op in operators:
        op_evaluations = [e for e in evaluations if e.operator_id == op.id]
        if op_evaluations:
            avg_quality = sum(e.quality_percent for e in op_evaluations) / len(op_evaluations)
            top_operators.append({
                "id": op.id,
                "name": op.full_name,
                "avg_quality": round(avg_quality, 1),
                "evaluations_count": len(op_evaluations)
            })
    
    top_operators.sort(key=lambda x: x["avg_quality"], reverse=True)
    
    return JSONResponse({
        "project_id": project_id,
        "project_name": project.name,
        "period_days": days,
        "total_evaluations": len(evaluations),
        "total_operators": len(operators),
        "avg_quality": round(sum(e.quality_percent for e in evaluations) / len(evaluations), 1) if evaluations else 0,
        "blocks_stats": blocks_stats,
        "top_operators": top_operators[:10]
    })

# ============== API ДЛЯ ДЕТАЛЬНОЙ СТАТИСТИКИ ОПЕРАТОРА ==============

@app.get("/api/operator/{operator_id}/detailed-stats")
async def get_operator_detailed_stats(
    operator_id: int,
    days: int = Query(30, description="Количество дней для анализа"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получение детальной статистики по оператору"""
    operator = db.query(Operator).filter(Operator.id == operator_id).first()
    if not operator:
        raise HTTPException(status_code=404, detail="Оператор не найден")
    
    # Получаем оценки за период
    cutoff_date = date.today() - timedelta(days=days)
    evaluations = db.query(Evaluation).filter(
        Evaluation.operator_id == operator_id,
        Evaluation.call_date >= cutoff_date
    ).order_by(Evaluation.call_date).all()
    
    # Статистика по дням
    daily_stats = []
    current = cutoff_date
    while current <= date.today():
        day_evaluations = [e for e in evaluations if e.call_date == current]
        if day_evaluations:
            avg_quality = sum(e.quality_percent for e in day_evaluations) / len(day_evaluations)
            daily_stats.append({
                "date": current.strftime("%Y-%m-%d"),
                "label": current.strftime("%d.%m"),
                "count": len(day_evaluations),
                "avg_quality": round(avg_quality, 1)
            })
        current += timedelta(days=1)
    
    # Статистика по блокам
    blocks_stats = []
    blocks = db.query(MetricBlock).filter(
        MetricBlock.project_id == operator.project_id,
        MetricBlock.is_active == True
    ).all()
    
    for block in blocks:
        block_results = []
        for eval_obj in evaluations:
            for br in eval_obj.block_results:
                if br.block_id == block.id:
                    block_results.append(br)
        
        if block_results:
            avg_percent = sum(r.percent for r in block_results) / len(block_results)
            resets_count = sum(1 for r in block_results if r.was_reset)
        else:
            avg_percent = 0
            resets_count = 0
        
        blocks_stats.append({
            "id": block.id,
            "name": block.name,
            "avg_percent": round(avg_percent, 1),
            "evaluations_count": len(block_results),
            "resets_count": resets_count
        })
    
    return JSONResponse({
        "operator_id": operator_id,
        "operator_name": operator.full_name,
        "period_days": days,
        "total_evaluations": len(evaluations),
        "avg_quality": round(operator.avg_quality, 1) if operator.avg_quality else 0,
        "daily_stats": daily_stats,
        "blocks_stats": blocks_stats
    })

# ============== ПРОЕКТЫ ==============

@app.get("/projects", response_class=HTMLResponse)
async def projects_page(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Страница управления проектами"""
    projects = db.query(Project).filter(Project.is_active == True).all()
    
    projects_with_stats = []
    for project in projects:
        # Считаем метрики через блоки
        metrics = db.query(Metric).join(MetricBlock).filter(
            MetricBlock.project_id == project.id,
            Metric.is_active == True,
            MetricBlock.is_active == True
        ).count()
        
        operators_count = db.query(Operator).filter(
            Operator.project_id == project.id,
            Operator.is_active == True
        ).count()
        
        stats = get_project_stats(project.id, period_days=30)
        
        projects_with_stats.append({
            "id": project.id,
            "name": project.name,
            "description": project.description,
            "metrics_count": metrics,
            "operators_count": operators_count,
            "avg_quality": stats['avg_quality'] if stats else 0,
            "evaluations_count": stats['evaluations_count'] if stats else 0
        })
    
    # Количество активных тикетов (новые + в работе)
    tickets_count = get_active_tickets_count(db)
    
    return templates.TemplateResponse(
        "projects.html",
        {
            "request": request,
            "user": current_user,
            "projects": projects_with_stats,
            "tickets_count": tickets_count,
            "date": date
        }
    )

@app.get("/project/{project_id}", response_class=HTMLResponse)
async def project_detail(
    request: Request,
    project_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Детальная страница проекта с метриками и операторами"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Проект не найден")
    
    # Получаем метрики через блоки
    metrics = db.query(Metric).join(MetricBlock).filter(
        MetricBlock.project_id == project_id,
        Metric.is_active == True,
        MetricBlock.is_active == True
    ).order_by(MetricBlock.display_order, Metric.display_order).all()
    
    operators = db.query(Operator).filter(
        Operator.project_id == project_id,
        Operator.is_active == True
    ).all()
    
    stats = get_project_stats(project_id, period_days=30)
    weekly = get_weekly_dynamics(project_id=project_id)
    monthly = get_monthly_dynamics(project_id=project_id)
    
    # Получаем все блоки проекта
    blocks = db.query(MetricBlock).filter(
        MetricBlock.project_id == project_id,
        MetricBlock.is_active == True
    ).order_by(MetricBlock.display_order).all()
    
    # Получаем все проекты для полосы проектов
    all_projects = db.query(Project).filter(Project.is_active == True).all()
    
    # Количество активных тикетов (новые + в работе)
    tickets_count = get_active_tickets_count(db)
    
    # Собираем статистику по каждому блоку за 30 дней
    cutoff_date = date.today() - timedelta(days=30)
    block_stats = []
    total_evaluations = 0
    
    for block in blocks:
        # Получаем все результаты оценок по этому блоку
        block_results = db.query(EvaluationBlockResult).join(
            Evaluation, Evaluation.id == EvaluationBlockResult.evaluation_id
        ).filter(
            EvaluationBlockResult.block_id == block.id,
            Evaluation.call_date >= cutoff_date
        ).all()
        
        if block_results:
            # Используем поле percent из EvaluationBlockResult
            total_percent = sum(r.percent for r in block_results)
            block_percent = total_percent / len(block_results)
            
            # Считаем успешные оценки (где не было обнуления)
            successful = sum(1 for r in block_results if not r.was_reset)
            total = len(block_results)
            resets = sum(1 for r in block_results if r.was_reset)
            
            total_evaluations += total
        else:
            block_percent = 0
            successful = 0
            total = 0
            resets = 0
        
        block_stats.append({
            "id": block.id,
            "name": block.name,
            "percent": block_percent,
            "successful": successful,
            "total": total,
            "resets": resets,
            "evaluations": total
        })
    
    return templates.TemplateResponse(
        "project_detail.html",
        {
            "request": request,
            "user": current_user,
            "project": project,
            "metrics": metrics,
            "operators": operators,
            "stats": stats,
            "weekly": weekly,
            "monthly": monthly,
            "blocks": blocks,
            "block_stats": block_stats,
            "total_evaluations": total_evaluations,
            "projects": all_projects,
            "tickets_count": tickets_count,
            "date": date
        }
    )

@app.get("/api/block/{block_id}/dynamics")
async def get_block_dynamics(
    block_id: int,
    start_date: str = Query(..., description="Начальная дата в формате ГГГГ-ММ-ДД"),
    end_date: str = Query(..., description="Конечная дата в формате ГГГГ-ММ-ДД"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получение динамики качества по конкретному блоку"""
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d").date()
        end = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Неверный формат даты")
    
    if start > end:
        raise HTTPException(status_code=400, detail="Начальная дата не может быть позже конечной")
    
    # Проверяем существование блока
    block = db.query(MetricBlock).filter(MetricBlock.id == block_id).first()
    if not block:
        raise HTTPException(status_code=404, detail="Блок не найден")
    
    # Получаем все результаты по блоку за период
    block_results = db.query(EvaluationBlockResult).join(
        Evaluation, Evaluation.id == EvaluationBlockResult.evaluation_id
    ).filter(
        EvaluationBlockResult.block_id == block_id,
        Evaluation.call_date >= start,
        Evaluation.call_date <= end
    ).order_by(Evaluation.call_date).all()
    
    # Группируем по дням
    daily_data = {}
    for result in block_results:
        eval_date = result.evaluation.call_date
        date_str = eval_date.strftime("%Y-%m-%d")
        
        if date_str not in daily_data:
            daily_data[date_str] = {
                "date": date_str,
                "total_percent": 0,
                "count": 0,
                "resets": 0
            }
        
        daily_data[date_str]["total_percent"] += result.percent
        daily_data[date_str]["count"] += 1
        if result.was_reset:
            daily_data[date_str]["resets"] += 1
    
    # Формируем результат
    result = []
    current = start
    while current <= end:
        date_str = current.strftime("%Y-%m-%d")
        if date_str in daily_data:
            data = daily_data[date_str]
            avg_percent = data["total_percent"] / data["count"] if data["count"] > 0 else 0
            
            result.append({
                "date": date_str,
                "label": current.strftime("%d.%m"),
                "percent": round(avg_percent, 1),
                "count": data["count"],
                "resets": data["resets"]
            })
        else:
            result.append({
                "date": date_str,
                "label": current.strftime("%d.%m"),
                "percent": 0,
                "count": 0,
                "resets": 0
            })
        current += timedelta(days=1)
    
    return JSONResponse(result)

@app.post("/api/project/create")
async def create_project(
    request: Request,
    current_user: User = Depends(require_role(['admin'])),
    db: Session = Depends(get_db)
):
    """Создание нового проекта (только админ)"""
    data = await request.json()
    
    project = Project(
        name=data.get("name"),
        description=data.get("description", "")
    )
    db.add(project)
    db.commit()
    db.refresh(project)
    
    # Создаем блок по умолчанию для нового проекта
    default_block = MetricBlock(
        project_id=project.id,
        name="Основные метрики",
        display_order=0,
        is_active=True
    )
    db.add(default_block)
    db.commit()
    
    return JSONResponse({
        "status": "success",
        "project_id": project.id,
        "message": f"Проект '{project.name}' создан"
    })

# ============== ОЦЕНКИ ЗВОНКОВ ==============

@app.get("/evaluate/{operator_id}", response_class=HTMLResponse)
async def evaluate_page(
    request: Request,
    operator_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Страница оценки звонка оператора"""
    if current_user.role not in ['admin', 'controller']:
        raise HTTPException(status_code=403, detail="Только контролеры могут оценивать звонки")
    
    operator = db.query(Operator).filter(Operator.id == operator_id).first()
    if not operator:
        raise HTTPException(status_code=404, detail="Оператор не найден")
    
    # Получаем метрики через блоки проекта
    metrics = db.query(Metric).join(MetricBlock).filter(
        MetricBlock.project_id == operator.project_id,
        Metric.is_active == True,
        MetricBlock.is_active == True
    ).order_by(MetricBlock.display_order, Metric.display_order).all()
    
    # Получаем блоки для группировки
    blocks = db.query(MetricBlock).filter(
        MetricBlock.project_id == operator.project_id,
        MetricBlock.is_active == True
    ).order_by(MetricBlock.display_order).all()
    
    # Получаем темы проекта
    topics = db.query(Topic).filter(
        Topic.project_id == operator.project_id,
        Topic.is_active == True
    ).order_by(Topic.display_order).all()
    
    # Получаем все проекты для полосы проектов
    projects = db.query(Project).filter(Project.is_active == True).all()
    
    # Количество активных тикетов (новые + в работе)
    tickets_count = get_active_tickets_count(db)
    
    return templates.TemplateResponse(
        "evaluate.html",
        {
            "request": request,
            "user": current_user,
            "operator": operator,
            "metrics": metrics,
            "blocks": blocks,
            "topics": topics,
            "projects": projects,
            "tickets_count": tickets_count,
            "today": date.today().strftime("%Y-%m-%d"),
            "date": date
        }
    )

@app.post("/api/evaluation/submit")
async def submit_evaluation(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Отправка оценки звонка"""
    # Проверка прав доступа
    if current_user.role not in ['admin', 'controller']:
        raise HTTPException(
            status_code=403, 
            detail="Только контролеры и администраторы могут оценивать звонки"
        )
    
    # Получаем и парсим данные
    try:
        data = await request.json()
        print(f"\n{'='*60}")
        print(f"📥 ПОЛУЧЕНЫ ДАННЫЕ ОЦЕНКИ:")
        print(f"{'='*60}")
        print(f"Operator ID: {data.get('operator_id')}")
        print(f"Call Date: {data.get('call_date')}")
        print(f"Evaluation Date: {data.get('evaluation_date')}")
        print(f"Call Duration: {data.get('call_duration')} секунд")
        print(f"Call Link: {data.get('call_link')}")
        print(f"Topic ID: {data.get('topic_id')}")
        print(f"Create Ticket: {data.get('create_ticket')}")
        print(f"Metrics count: {len(data.get('metrics', []))}")
        if 'optional_blocks' in data:
            print(f"Optional blocks: {data.get('optional_blocks')}")
        print(f"{'='*60}\n")
    except Exception as e:
        print(f"❌ Ошибка парсинга JSON: {e}")
        raise HTTPException(
            status_code=400, 
            detail=f"Неверный формат данных: {str(e)}"
        )
    
    # Валидация обязательных полей
    operator_id = data.get("operator_id")
    if not operator_id:
        raise HTTPException(status_code=400, detail="operator_id обязателен")
    
    call_date = data.get("call_date")
    if not call_date:
        raise HTTPException(status_code=400, detail="call_date обязателен")
    
    # Преобразование дат
    try:
        call_date_obj = datetime.strptime(call_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Неверный формат даты звонка (требуется ГГГГ-ММ-ДД)")
    
    evaluation_date = data.get("evaluation_date")
    try:
        evaluation_date_obj = datetime.strptime(evaluation_date, "%Y-%m-%d").date() if evaluation_date else date.today()
    except ValueError:
        evaluation_date_obj = date.today()
    
    # Получение дополнительных полей
    call_link = data.get("call_link", "")
    comment = data.get("comment", "")
    topic_id = data.get("topic_id")
    metrics_data = data.get("metrics", [])
    
    # ПОЛУЧЕНИЕ ДЛИТЕЛЬНОСТИ ЗВОНКА
    call_duration = data.get("call_duration")
    
    # Преобразуем в int, если пришло как строка
    if call_duration is not None:
        try:
            call_duration = int(call_duration)
            if call_duration < 0:
                call_duration = 0
                print("⚠️ Длительность не может быть отрицательной, установлено 0")
        except (ValueError, TypeError):
            call_duration = None
            print("⚠️ Неверный формат длительности, значение игнорируется")
    
    # Данные для тикета
    create_ticket = data.get("create_ticket", False)
    ticket_topic = data.get("ticket_topic", "evaluation")
    ticket_description = data.get("ticket_description", "")
    
    # Данные об опциональных блоках
    optional_blocks_data = data.get("optional_blocks", [])
    
    # Получаем оператора
    operator = db.query(Operator).filter(Operator.id == operator_id).first()
    if not operator:
        raise HTTPException(status_code=404, detail="Оператор не найден")
    
    # Создаем оценку
    evaluation = Evaluation(
        operator_id=operator_id,
        evaluator_id=current_user.id,
        call_link=call_link,
        call_date=call_date_obj,
        call_duration=call_duration,
        evaluation_date=evaluation_date_obj,
        comment=comment,
        topic_id=topic_id if topic_id else None
    )
    
    try:
        # Сохраняем оценку
        db.add(evaluation)
        db.flush()
        
        # Добавляем метрики
        metrics_added = 0
        for m in metrics_data:
            metric = db.query(Metric).filter(Metric.id == m["metric_id"]).first()
            if not metric:
                continue
            
            metric_comment = m.get("comment", "")
            
            eval_metric = EvaluationMetric(
                evaluation_id=evaluation.id,
                metric_id=m["metric_id"],
                earned_score=m.get("earned_score"),
                max_score=metric.max_score,
                comment=metric_comment,
                is_not_evaluated=m.get("is_not_evaluated", False)
            )
            db.add(eval_metric)
            metrics_added += 1
        
        # Сохраняем метрики
        db.flush()
        
        # Пересчитываем проценты
        final_quality = calculate_evaluation_scores(evaluation, db)
        
        # Обновляем статистику оператора
        all_evaluations = db.query(Evaluation).filter(
            Evaluation.operator_id == operator.id
        ).all()
        
        operator.total_evaluations = len(all_evaluations)
        if all_evaluations:
            operator.avg_quality = sum(e.quality_percent for e in all_evaluations) / len(all_evaluations)
        
        # Создаем тикет если нужно
        ticket_created = False
        if create_ticket:
            try:
                # Рассчитываем SLA (2 рабочих дня)
                sla_deadline = calculate_sla_deadline(date.today())
                
                ticket_title = f"Проработка по оценке #{evaluation.id}"
                if ticket_topic == "general":
                    ticket_title = f"Общая проработка: {operator.full_name}"
                
                ticket = Ticket(
                    title=ticket_title,
                    description=ticket_description or f"Требуется проработка по оценке звонка оператора {operator.full_name}",
                    evaluation_id=evaluation.id,
                    operator_id=operator_id,
                    project_id=operator.project_id,
                    created_by=current_user.id,
                    topic=ticket_topic,
                    status=TicketStatus.NEW.value,
                    sla_deadline=sla_deadline,
                    created_at=datetime.utcnow()
                )
                db.add(ticket)
                db.flush()
                
                # Добавляем запись в историю
                history = TicketHistory(
                    ticket_id=ticket.id,
                    user_id=current_user.id,
                    action="created",
                    comment="Тикет создан по результатам оценки"
                )
                db.add(history)
                
                ticket_created = True
            except Exception as e:
                print(f"⚠️ Ошибка при создании тикета: {e}")
        
        # Сохраняем все изменения
        db.commit()
        
        return JSONResponse({
            "status": "success",
            "evaluation_id": evaluation.id,
            "quality_percent": round(evaluation.quality_percent, 1),
            "ticket_created": ticket_created,
            "call_duration": evaluation.call_duration
        })
        
    except Exception as e:
        db.rollback()
        print(f"\n❌ КРИТИЧЕСКАЯ ОШИБКА ПРИ СОХРАНЕНИИ: {str(e)}")
        import traceback
        traceback.print_exc()
        
        raise HTTPException(
            status_code=500, 
            detail=f"Ошибка сохранения оценки: {str(e)}"
        )

# ============== УПРАВЛЕНИЕ МЕТРИКАМИ ==============

# ============== УПРАВЛЕНИЕ МЕТРИКАМИ (ОБНОВЛЕННОЕ) ==============

@app.get("/project/{project_id}/metrics", response_class=HTMLResponse)
async def project_metrics_page(
    request: Request,
    project_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Страница управления метриками проекта"""
    if current_user.role not in ['admin', 'controller']:
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Проект не найден")
    
    # Получаем метрики через блоки
    metrics = db.query(Metric).join(MetricBlock).filter(
        MetricBlock.project_id == project_id
    ).order_by(MetricBlock.display_order, Metric.display_order).all()
    
    # Получаем все блоки проекта
    blocks = db.query(MetricBlock).filter(
        MetricBlock.project_id == project_id
    ).order_by(MetricBlock.display_order).all()
    
    # Получаем все проекты для полосы проектов
    all_projects = db.query(Project).filter(Project.is_active == True).all()
    
    # Количество активных тикетов (новые + в работе)
    tickets_count = get_active_tickets_count(db)
    
    return templates.TemplateResponse(
        "metrics.html",
        {
            "request": request,
            "user": current_user,
            "project": project,
            "metrics": metrics,
            "blocks": blocks,
            "projects": all_projects,
            "tickets_count": tickets_count,
            "date": date
        }
    )


@app.post("/api/project/{project_id}/metrics")
async def add_metric(
    project_id: int,
    request: Request,
    current_user: User = Depends(require_role(['admin', 'controller'])),
    db: Session = Depends(get_db)
):
    """Добавление новой метрики (с поддержкой resets_all)"""
    data = await request.json()
    
    block_id = data.get("block_id")
    if not block_id:
        block = db.query(MetricBlock).filter(
            MetricBlock.project_id == project_id,
            MetricBlock.is_active == True
        ).first()
        if not block:
            raise HTTPException(status_code=400, detail="Сначала создайте блок метрик")
        block_id = block.id
    
    max_order = db.query(func.max(Metric.display_order)).filter(
        Metric.block_id == block_id
    ).scalar() or 0
    
    max_score = data.get("max_score", 1)
    try:
        max_score = int(max_score)
    except (ValueError, TypeError):
        max_score = 1
    
    allowed_scores = [0, 1, 2, 3, 5, 10]
    if max_score not in allowed_scores:
        raise HTTPException(
            status_code=400, 
            detail=f"Максимальный балл должен быть одним из: {allowed_scores}"
        )
    
    metric = Metric(
        block_id=block_id,
        name=data.get("name"),
        description=data.get("description", ""),
        max_score=max_score,
        is_critical=data.get("is_critical", False),
        is_global_critical=data.get("is_global_critical", False),
        allow_na=data.get("allow_na", True),
        resets_block=data.get("resets_block", False),
        resets_all=data.get("resets_all", False),  # НОВОЕ ПОЛЕ
        penalty_type=data.get("penalty_type") if data.get("is_critical") else None,
        penalty_value=data.get("penalty_value") if data.get("is_critical") else None,
        display_order=max_order + 1
    )
    db.add(metric)
    db.commit()
    db.refresh(metric)
    
    return JSONResponse({
        "status": "success",
        "metric": {
            "id": metric.id,
            "name": metric.name,
            "max_score": metric.max_score,
            "block_id": metric.block_id,
            "is_critical": metric.is_critical,
            "is_global_critical": metric.is_global_critical,
            "allow_na": metric.allow_na,
            "resets_block": metric.resets_block,
            "resets_all": metric.resets_all  # НОВОЕ ПОЛЕ
        }
    })


@app.post("/api/project/{project_id}/blocks")
async def create_metric_block(
    project_id: int,
    request: Request,
    current_user: User = Depends(require_role(['admin', 'controller'])),
    db: Session = Depends(get_db)
):
    """Создание нового блока метрик"""
    data = await request.json()
    
    max_order = db.query(func.max(MetricBlock.display_order)).filter(
        MetricBlock.project_id == project_id
    ).scalar() or 0
    
    block = MetricBlock(
        project_id=project_id,
        name=data.get("name"),
        description=data.get("description", ""),
        display_order=max_order + 1
    )
    db.add(block)
    db.commit()
    db.refresh(block)
    
    return JSONResponse({
        "status": "success",
        "block": {
            "id": block.id,
            "name": block.name,
            "description": block.description,
            "display_order": block.display_order
        }
    })


@app.put("/api/metrics/{metric_id}")
async def update_metric(
    metric_id: int,
    request: Request,
    current_user: User = Depends(require_role(['admin', 'controller'])),
    db: Session = Depends(get_db)
):
    """Обновление метрики (с поддержкой resets_all)"""
    metric = db.query(Metric).filter(Metric.id == metric_id).first()
    if not metric:
        raise HTTPException(status_code=404, detail="Метрика не найдена")
    
    data = await request.json()
    
    # Обновляем поля
    if "name" in data:
        metric.name = data["name"]
    
    if "description" in data:
        metric.description = data["description"]
    
    if "max_score" in data:
        try:
            max_score = int(data["max_score"])
            allowed_scores = [0, 1, 2, 3, 5, 10]
            if max_score not in allowed_scores:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Максимальный балл должен быть одним из: {allowed_scores}"
                )
            metric.max_score = max_score
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="Неверный формат максимального балла")
    
    if "is_critical" in data:
        metric.is_critical = data["is_critical"]
    
    if "is_global_critical" in data:
        metric.is_global_critical = data["is_global_critical"]
    
    if "allow_na" in data:
        metric.allow_na = data["allow_na"]
    
    if "resets_block" in data:
        metric.resets_block = data["resets_block"]
    
    if "resets_all" in data:  # НОВОЕ ПОЛЕ
        metric.resets_all = data["resets_all"]
    
    if "penalty_type" in data:
        metric.penalty_type = data["penalty_type"]
    
    if "penalty_value" in data:
        metric.penalty_value = data["penalty_value"]
    
    if "is_active" in data:
        metric.is_active = data["is_active"]
    
    db.commit()
    
    return JSONResponse({"status": "success"})


@app.delete("/api/metrics/{metric_id}")
async def delete_metric(
    metric_id: int,
    current_user: User = Depends(require_role(['admin', 'controller'])),
    db: Session = Depends(get_db)
):
    """Удаление метрики"""
    metric = db.query(Metric).filter(Metric.id == metric_id).first()
    if not metric:
        raise HTTPException(status_code=404, detail="Метрика не найдена")
    
    # Проверяем, используется ли метрика в оценках
    evaluations_count = db.query(EvaluationMetric).filter(EvaluationMetric.metric_id == metric_id).count()
    if evaluations_count > 0:
        # Если есть оценки, просто деактивируем
        metric.is_active = False
        db.commit()
        return JSONResponse({
            "status": "success", 
            "message": f"Метрика деактивирована, так как используется в {evaluations_count} оценках"
        })
    else:
        # Если нет оценок, удаляем полностью
        db.delete(metric)
        db.commit()
        return JSONResponse({"status": "success", "message": "Метрика удалена"})


@app.delete("/api/blocks/{block_id}")
async def delete_block(
    block_id: int,
    current_user: User = Depends(require_role(['admin', 'controller'])),
    db: Session = Depends(get_db)
):
    """Удаление блока метрик"""
    block = db.query(MetricBlock).filter(MetricBlock.id == block_id).first()
    if not block:
        raise HTTPException(status_code=404, detail="Блок не найден")
    
    # Проверяем, есть ли метрики в этом блоке
    metrics_count = db.query(Metric).filter(Metric.block_id == block_id).count()
    if metrics_count > 0:
        # Если есть метрики, просто деактивируем блок
        block.is_active = False
        db.commit()
        return JSONResponse({
            "status": "success",
            "message": f"Блок деактивирован, так как содержит {metrics_count} метрик"
        })
    else:
        # Если нет метрик, удаляем полностью
        db.delete(block)
        db.commit()
        return JSONResponse({
            "status": "success",
            "message": "Блок удален"
        })


@app.put("/api/blocks/{block_id}")
async def update_block(
    block_id: int,
    request: Request,
    current_user: User = Depends(require_role(['admin', 'controller'])),
    db: Session = Depends(get_db)
):
    """Обновление блока метрик"""
    block = db.query(MetricBlock).filter(MetricBlock.id == block_id).first()
    if not block:
        raise HTTPException(status_code=404, detail="Блок не найден")
    
    data = await request.json()
    
    # Обновляем поля
    if "name" in data:
        block.name = data["name"]
    if "description" in data:
        block.description = data["description"]
    
    db.commit()
    
    return JSONResponse({
        "status": "success",
        "block": {
            "id": block.id,
            "name": block.name,
            "description": block.description,
            "display_order": block.display_order
        }
    })


@app.get("/api/projects/{project_id}/metrics")
async def get_project_metrics(
    project_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получение метрик проекта (с полем resets_all)"""
    metrics = db.query(Metric).join(MetricBlock).filter(
        MetricBlock.project_id == project_id,
        Metric.is_active == True,
        MetricBlock.is_active == True
    ).order_by(MetricBlock.display_order, Metric.display_order).all()
    
    result = []
    for metric in metrics:
        result.append({
            "id": metric.id,
            "name": metric.name,
            "description": metric.description,
            "max_score": metric.max_score,
            "is_critical": metric.is_critical,
            "is_global_critical": metric.is_global_critical,
            "allow_na": metric.allow_na,
            "resets_block": metric.resets_block,
            "resets_all": metric.resets_all,  # НОВОЕ ПОЛЕ
            "penalty_type": metric.penalty_type,
            "penalty_value": metric.penalty_value,
            "block_id": metric.block_id,
            "block_name": metric.block.name,
            "display_order": metric.display_order
        })
    
    return JSONResponse(result)


@app.get("/api/projects/{project_id}/blocks")
async def get_project_blocks(
    project_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получение блоков метрик проекта"""
    blocks = db.query(MetricBlock).filter(
        MetricBlock.project_id == project_id,
        MetricBlock.is_active == True
    ).order_by(MetricBlock.display_order).all()
    
    result = []
    for block in blocks:
        metrics_count = db.query(Metric).filter(
            Metric.block_id == block.id,
            Metric.is_active == True
        ).count()
        
        result.append({
            "id": block.id,
            "name": block.name,
            "description": block.description,
            "display_order": block.display_order,
            "metrics_count": metrics_count
        })
    
    return JSONResponse(result)
# ============== ОТЧЕТЫ И СТАТИСТИКА ==============

@app.get("/reports", response_class=HTMLResponse)
async def reports_page(
    request: Request,
    project_id: Optional[int] = None,
    operator_id: Optional[int] = None,
    period: str = Query("month", regex="^(day|week|month|quarter|custom)$"),
    start_date: Optional[str] = Query(None, description="Начальная дата для кастомного периода"),
    end_date: Optional[str] = Query(None, description="Конечная дата для кастомного периода"),
    metric_id: Optional[int] = Query(None, description="ID метрики для фильтрации"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Страница с отчетами и графиками"""
    projects = db.query(Project).filter(Project.is_active == True).all()
    
    # Получаем операторов для фильтра
    operators_query = db.query(Operator).filter(Operator.is_active == True)
    if project_id:
        operators_query = operators_query.filter(Operator.project_id == project_id)
    operators = operators_query.all()
    
    # Для статистики используем стандартный период 30 дней
    if project_id:
        stats = get_project_stats(project_id, period_days=30)
    elif operator_id:
        stats = get_operator_stats(operator_id, period_days=30)
    else:
        stats = None
    
    # Топ операторов (всегда общий)
    top_operators = db.query(
        Operator.id,
        Operator.full_name,
        Operator.project_id,
        Project.name.label('project_name'),
        func.avg(Evaluation.quality_percent).label('avg_quality'),
        func.count(Evaluation.id).label('evaluations_count')
    ).join(Evaluation).join(Project).group_by(Operator.id).order_by(
        desc('avg_quality')
    ).limit(10).all()
    
    # Количество активных тикетов (новые + в работе)
    tickets_count = get_active_tickets_count(db)
    
    return templates.TemplateResponse(
        "reports.html",
        {
            "request": request,
            "user": current_user,
            "projects": projects,
            "operators": operators,
            "selected_project": project_id,
            "selected_operator": operator_id,
            "period": period,
            "start_date": start_date,
            "end_date": end_date,
            "metric_id": metric_id,
            "stats": stats,
            "top_operators": top_operators,
            "tickets_count": tickets_count,
            "date": date
        }
    )

@app.get("/api/statistics/operator/{operator_id}")
async def api_operator_stats(
    operator_id: int,
    period: int = 30,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """API для получения статистики оператора"""
    stats = get_operator_stats(operator_id, period)
    if not stats:
        raise HTTPException(status_code=404, detail="Нет данных")
    
    weekly = get_weekly_dynamics(operator_id=operator_id)
    monthly = get_monthly_dynamics(operator_id=operator_id)
    
    return JSONResponse({
        "stats": stats,
        "weekly": weekly,
        "monthly": monthly
    })

@app.get("/api/statistics/project/{project_id}")
async def api_project_stats(
    project_id: int,
    period: int = 30,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """API для получения статистики проекта"""
    stats = get_project_stats(project_id, period)
    if not stats:
        raise HTTPException(status_code=404, detail="Нет данных")
    
    weekly = get_weekly_dynamics(project_id=project_id)
    monthly = get_monthly_dynamics(project_id=project_id)
    
    return JSONResponse({
        "stats": stats,
        "weekly": weekly,
        "monthly": monthly
    })

# ============== API ДЛЯ ДИНАМИКИ ==============

@app.get("/api/operator/{operator_id}/dynamics")
async def get_operator_dynamics(
    operator_id: int,
    start_date: str = Query(..., description="Начальная дата в формате ГГГГ-ММ-ДД"),
    end_date: str = Query(..., description="Конечная дата в формате ГГГГ-ММ-ДД"),
    metric_id: Optional[int] = Query(None, description="ID метрики для фильтрации"),
    interval: str = Query("week", regex="^(day|week|month)$"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получение динамики качества оператора с фильтрацией"""
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d").date()
        end = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Неверный формат даты")
    
    if start > end:
        raise HTTPException(status_code=400, detail="Начальная дата не может быть позже конечной")
    
    operator = db.query(Operator).filter(Operator.id == operator_id).first()
    if not operator:
        raise HTTPException(status_code=404, detail="Оператор не найден")
    
    # Получаем все оценки за период
    evaluations = db.query(Evaluation).filter(
        Evaluation.operator_id == operator_id,
        Evaluation.call_date >= start,
        Evaluation.call_date <= end
    ).order_by(Evaluation.call_date).all()
    
    # Определяем шаг интервала
    if interval == "day":
        delta = timedelta(days=1)
    elif interval == "week":
        delta = timedelta(days=7)
    else:  # month
        delta = timedelta(days=30)
    
    # Группируем по интервалам
    result = []
    current = start
    
    while current <= end:
        interval_end = current + delta - timedelta(days=1)
        if interval_end > end:
            interval_end = end
        
        # Фильтруем оценки за интервал
        interval_evals = [e for e in evaluations if current <= e.call_date <= interval_end]
        
        if metric_id and metric_id != 0:
            # Фильтр по конкретной метрике
            total_percent = 0
            count = 0
            
            for eval_obj in interval_evals:
                for metric in eval_obj.metrics:
                    if metric.metric_id == metric_id and not metric.is_not_evaluated and metric.earned_score is not None:
                        metric_percent = (metric.earned_score / metric.max_score) * 100
                        total_percent += metric_percent
                        count += 1
            
            avg_quality = total_percent / count if count > 0 else 0
        else:
            # Общее качество
            avg_quality = sum(e.quality_percent for e in interval_evals) / len(interval_evals) if interval_evals else 0
        
        # Формируем название интервала
        if interval == "day":
            label = current.strftime("%d.%m")
        elif interval == "week":
            label = f"{current.strftime('%d.%m')}-{interval_end.strftime('%d.%m')}"
        else:  # month
            label = current.strftime("%b %Y")
        
        result.append({
            "date": current.strftime("%Y-%m-%d"),
            "label": label,
            "quality": round(avg_quality, 1),
            "count": len(interval_evals)
        })
        
        current = interval_end + timedelta(days=1)
    
    return JSONResponse(result)


@app.get("/api/project/{project_id}/dynamics")
async def get_project_dynamics(
    project_id: int,
    start_date: str = Query(..., description="Начальная дата в формате ГГГГ-ММ-ДД"),
    end_date: str = Query(..., description="Конечная дата в формате ГГГГ-ММ-ДД"),
    interval: str = Query("week", regex="^(day|week|month)$"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получение динамики качества проекта"""
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d").date()
        end = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Неверный формат даты")
    
    if start > end:
        raise HTTPException(status_code=400, detail="Начальная дата не может быть позже конечной")
    
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Проект не найден")
    
    # Получаем операторов проекта
    operators = db.query(Operator.id).filter(
        Operator.project_id == project_id,
        Operator.is_active == True
    ).subquery()
    
    # Получаем все оценки за период
    evaluations = db.query(Evaluation).filter(
        Evaluation.operator_id.in_(operators),
        Evaluation.call_date >= start,
        Evaluation.call_date <= end
    ).order_by(Evaluation.call_date).all()
    
    # Определяем шаг интервала
    if interval == "day":
        delta = timedelta(days=1)
    elif interval == "week":
        delta = timedelta(days=7)
    else:  # month
        delta = timedelta(days=30)
    
    # Группируем по интервалам
    result = []
    current = start
    
    while current <= end:
        interval_end = current + delta - timedelta(days=1)
        if interval_end > end:
            interval_end = end
        
        interval_evals = [e for e in evaluations if current <= e.call_date <= interval_end]
        avg_quality = sum(e.quality_percent for e in interval_evals) / len(interval_evals) if interval_evals else 0
        
        if interval == "day":
            label = current.strftime("%d.%m")
        elif interval == "week":
            label = f"{current.strftime('%d.%m')}-{interval_end.strftime('%d.%m')}"
        else:
            label = current.strftime("%b %Y")
        
        result.append({
            "date": current.strftime("%Y-%m-%d"),
            "label": label,
            "quality": round(avg_quality, 1),
            "count": len(interval_evals)
        })
        
        current = interval_end + timedelta(days=1)
    
    return JSONResponse(result)

# ============== ТИКЕТЫ ==============

@app.get("/tickets", response_class=HTMLResponse)
async def tickets_page(
    request: Request,
    status: Optional[str] = Query(None, description="Фильтр по статусу"),
    project_id: Optional[str] = Query(None, description="Фильтр по проекту (ID)"),
    start_date: Optional[str] = Query(None, description="Начальная дата"),
    end_date: Optional[str] = Query(None, description="Конечная дата"),
    date_type: Optional[str] = Query("created", description="Тип даты: created или completed"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Страница со списком тикетов"""
    projects = db.query(Project).filter(Project.is_active == True).all()
    
    # Преобразуем project_id из строки в число, если оно не пустое
    project_id_int = None
    if project_id and project_id.strip() and project_id.strip().isdigit():
        project_id_int = int(project_id.strip())
    
    # Получаем доступные проекты для пользователя
    user_projects = get_user_projects(current_user, db)
    
    # Базовый запрос
    query = db.query(Ticket).join(Operator).join(Project)
    
    # Фильтры по ролям
    if current_user.role == 'controller':
        if project_id_int:
            query = query.filter(Ticket.project_id == project_id_int)
    elif current_user.role == 'manager':
        if user_projects:
            query = query.filter(Ticket.project_id.in_(user_projects))
            if project_id_int and project_id_int in user_projects:
                query = query.filter(Ticket.project_id == project_id_int)
        else:
            query = query.filter(False)
    
    # Фильтр по статусу
    if status and status.strip():
        query = query.filter(Ticket.status == status.strip())
    
    # Фильтры по датам
    if start_date and start_date.strip():
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").date()
            if date_type == "created":
                query = query.filter(Ticket.created_at >= datetime.combine(start, datetime.min.time()))
            else:  # completed
                query = query.filter(Ticket.completed_at >= start)
        except ValueError:
            pass
    
    if end_date and end_date.strip():
        try:
            end = datetime.strptime(end_date, "%Y-%m-%d").date()
            if date_type == "created":
                query = query.filter(Ticket.created_at <= datetime.combine(end, datetime.max.time()))
            else:  # completed
                query = query.filter(Ticket.completed_at <= end)
        except ValueError:
            pass
    
    # Сортировка
    tickets = query.order_by(
        case(
            (Ticket.status == TicketStatus.NEW.value, 1),
            (Ticket.status == TicketStatus.IN_PROGRESS.value, 2),
            (Ticket.status == TicketStatus.COMPLETED.value, 3),
            else_=4
        ),
        Ticket.sla_deadline.asc(),
        Ticket.created_at.desc()
    ).all()
    
    # Статистика по тикетам (с учетом прав доступа)
    base_stats_query = db.query(Ticket)
    
    if current_user.role == 'manager' and user_projects:
        base_stats_query = base_stats_query.filter(Ticket.project_id.in_(user_projects))
    
    stats = {
        "new": base_stats_query.filter(Ticket.status == TicketStatus.NEW.value).count(),
        "in_progress": base_stats_query.filter(Ticket.status == TicketStatus.IN_PROGRESS.value).count(),
        "overdue": base_stats_query.filter(
            Ticket.status.in_([TicketStatus.NEW.value, TicketStatus.IN_PROGRESS.value]),
            Ticket.sla_deadline < date.today()
        ).count()
    }
    
    # Получаем все проекты для полосы проектов
    all_projects = db.query(Project).filter(Project.is_active == True).all()
    
    # Количество активных тикетов для бейджа
    tickets_count = get_active_tickets_count(db)
    
    return templates.TemplateResponse(
        "tickets.html",
        {
            "request": request,
            "user": current_user,
            "projects": all_projects,
            "tickets": tickets,
            "stats": stats,
            "selected_status": status if status and status.strip() else None,
            "selected_project": project_id_int,
            "start_date": start_date,
            "end_date": end_date,
            "date_type": date_type,
            "tickets_count": tickets_count,
            "date": date
        }
    )

@app.get("/api/tickets")
async def get_tickets(
    status: Optional[str] = Query(None),
    project_id: Optional[int] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """API для получения списка тикетов"""
    # Получаем доступные проекты для пользователя
    user_projects = get_user_projects(current_user, db)
    
    query = db.query(Ticket).join(Operator).join(Project)
    
    # Фильтры по ролям
    if current_user.role == 'controller':
        # Контролеры видят все тикеты
        pass
        
    elif current_user.role == 'manager':
        # Менеджеры видят тикеты только по своим проектам
        if user_projects:
            query = query.filter(Ticket.project_id.in_(user_projects))
        else:
            query = query.filter(False)
    
    if status:
        query = query.filter(Ticket.status == status)
    if project_id:
        if current_user.role == 'manager' and project_id not in user_projects:
            # Если менеджер запрашивает проект, к которому не имеет доступа
            return JSONResponse([])
        query = query.filter(Ticket.project_id == project_id)
    
    tickets = query.order_by(Ticket.created_at.desc()).all()
    
    result = []
    for ticket in tickets:
        result.append({
            "id": ticket.id,
            "title": ticket.title,
            "description": ticket.description,
            "operator": ticket.operator.full_name,
            "project": ticket.project.name,
            "status": ticket.status,
            "priority": ticket.priority,
            "topic": ticket.topic,
            "sla_deadline": ticket.sla_deadline.strftime("%d.%m.%Y") if ticket.sla_deadline else None,
            "created_at": ticket.created_at.strftime("%d.%m.%Y %H:%M"),
            "created_by": ticket.creator.full_name,
            "assigned_to": ticket.assignee.full_name if ticket.assignee else None
        })
    
    return JSONResponse(result)


@app.post("/api/tickets/{ticket_id}/take")
async def take_ticket(
    ticket_id: int,
    current_user: User = Depends(require_role(['admin'])),
    db: Session = Depends(get_db)
):
    """Взять тикет в работу (только админ)"""
    ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Тикет не найден")
    
    if ticket.status != TicketStatus.NEW.value:
        raise HTTPException(status_code=400, detail="Тикет уже в работе или завершен")
    
    ticket.status = TicketStatus.IN_PROGRESS.value
    ticket.assigned_to = current_user.id
    ticket.taken_at = datetime.utcnow()
    
    # Добавляем в историю
    history = TicketHistory(
        ticket_id=ticket.id,
        user_id=current_user.id,
        action="taken",
        comment="Тикет взят в работу"
    )
    db.add(history)
    
    db.commit()
    
    return JSONResponse({
        "status": "success",
        "message": "Тикет взят в работу"
    })


@app.post("/api/tickets/{ticket_id}/complete")
async def complete_ticket(
    ticket_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Завершить тикет (админ или тот, кто взял в работу)"""
    data = await request.json()
    
    ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Тикет не найден")
    
    # Проверяем права
    if current_user.role != 'admin' and ticket.assigned_to != current_user.id:
        raise HTTPException(status_code=403, detail="Нет прав на завершение этого тикета")
    
    if ticket.status != TicketStatus.IN_PROGRESS.value:
        raise HTTPException(status_code=400, detail="Тикет не в работе")
    
    # Обновляем поля тикета
    ticket.status = TicketStatus.COMPLETED.value
    ticket.completed_at = datetime.utcnow()
    ticket.meeting_link = data.get("meeting_link")
    ticket.work_comment = data.get("work_comment")
    
    if data.get("block_id"):
        ticket.block_id = data.get("block_id")
    if data.get("metric_id"):
        ticket.metric_id = data.get("metric_id")
    
    # Формируем комментарий для истории
    history_comment = data.get("work_comment") or data.get("comment", "Тикет завершен")
    
    # Добавляем в историю
    history = TicketHistory(
        ticket_id=ticket.id,
        user_id=current_user.id,
        action="completed",
        comment=history_comment
    )
    db.add(history)
    
    db.commit()
    
    return JSONResponse({
        "status": "success",
        "message": "Тикет завершен"
    })

@app.get("/api/tickets/{ticket_id}")
async def get_ticket(
    ticket_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получить детальную информацию по тикету"""
    ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Тикет не найден")
    
    # Проверяем права доступа
    user_projects = get_user_projects(current_user, db)
    if current_user.role == 'manager' and ticket.project_id not in user_projects:
        raise HTTPException(status_code=403, detail="Нет доступа к этому тикету")
    
    # Получаем блоки и метрики проекта для фильтров
    blocks = db.query(MetricBlock).filter(
        MetricBlock.project_id == ticket.project_id,
        MetricBlock.is_active == True
    ).order_by(MetricBlock.display_order).all()
    
    metrics = []
    if ticket.block_id:
        metrics = db.query(Metric).filter(
            Metric.block_id == ticket.block_id,
            Metric.is_active == True
        ).order_by(Metric.display_order).all()
    
    # История тикета
    history = db.query(TicketHistory).filter(
        TicketHistory.ticket_id == ticket.id
    ).order_by(TicketHistory.created_at.desc()).all()
    
    return JSONResponse({
        "ticket": {
            "id": ticket.id,
            "title": ticket.title,
            "description": ticket.description,
            "operator": {
                "id": ticket.operator.id,
                "name": ticket.operator.full_name
            },
            "project": {
                "id": ticket.project.id,
                "name": ticket.project.name
            },
            "evaluation_id": ticket.evaluation_id,
            "status": ticket.status,
            "priority": ticket.priority,
            "topic": ticket.topic,
            "sla_deadline": ticket.sla_deadline.strftime("%Y-%m-%d") if ticket.sla_deadline else None,
            "created_at": ticket.created_at.strftime("%Y-%m-%d %H:%M"),
            "taken_at": ticket.taken_at.strftime("%Y-%m-%d %H:%M") if ticket.taken_at else None,
            "completed_at": ticket.completed_at.strftime("%Y-%m-%d %H:%M") if ticket.completed_at else None,
            "created_by": ticket.creator.full_name,
            "assigned_to": ticket.assignee.full_name if ticket.assignee else None,
            "meeting_link": ticket.meeting_link,
            "work_comment": ticket.work_comment,
            "block_id": ticket.block_id,
            "metric_id": ticket.metric_id
        },
        "blocks": [
            {"id": b.id, "name": b.name} for b in blocks
        ],
        "metrics": [
            {"id": m.id, "name": m.name} for m in metrics
        ],
        "history": [
            {
                "action": h.action,
                "user": h.user.full_name,
                "comment": h.comment,
                "created_at": h.created_at.strftime("%d.%m.%Y %H:%M")
            } for h in history
        ]
    })

@app.get("/tickets/{ticket_id}", response_class=HTMLResponse)
async def ticket_detail(
    request: Request,
    ticket_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Детальная страница тикета"""
    # Получаем тикет
    ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Тикет не найден")
    
    # Проверяем права доступа
    user_projects = get_user_projects(current_user, db)
    if current_user.role == 'manager' and ticket.project_id not in user_projects:
        raise HTTPException(status_code=403, detail="Нет доступа к этому тикету")
    
    # Получаем блоки проекта для фильтра
    blocks = db.query(MetricBlock).filter(
        MetricBlock.project_id == ticket.project_id,
        MetricBlock.is_active == True
    ).order_by(MetricBlock.display_order).all()
    
    # Получаем метрики если выбран блок
    metrics = []
    if ticket.block_id:
        metrics = db.query(Metric).filter(
            Metric.block_id == ticket.block_id,
            Metric.is_active == True
        ).order_by(Metric.display_order).all()
    
    # Получаем историю тикета
    history = db.query(TicketHistory).filter(
        TicketHistory.ticket_id == ticket.id
    ).order_by(TicketHistory.created_at.desc()).all()
    
    # Получаем все проекты для полосы проектов
    projects = db.query(Project).filter(Project.is_active == True).all()
    
    # Количество активных тикетов для бейджа
    tickets_count = get_active_tickets_count(db)
    
    return templates.TemplateResponse(
        "ticket_detail.html",
        {
            "request": request,
            "user": current_user,
            "ticket": ticket,
            "blocks": blocks,
            "metrics": metrics,
            "history": history,
            "projects": projects,
            "tickets_count": tickets_count,
            "date": date
        }
    )

@app.get("/api/blocks/{block_id}/metrics")
async def get_block_metrics(
    block_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получение метрик для блока"""
    metrics = db.query(Metric).filter(
        Metric.block_id == block_id,
        Metric.is_active == True
    ).order_by(Metric.display_order).all()
    
    result = []
    for metric in metrics:
        result.append({
            "id": metric.id,
            "name": metric.name,
            "max_score": metric.max_score
        })
    
    return JSONResponse(result)

@app.put("/api/blocks/{block_id}")
async def update_block(
    block_id: int,
    request: Request,
    current_user: User = Depends(require_role(['admin', 'controller'])),
    db: Session = Depends(get_db)
):
    """Обновление блока метрик"""
    block = db.query(MetricBlock).filter(MetricBlock.id == block_id).first()
    if not block:
        raise HTTPException(status_code=404, detail="Блок не найден")
    
    data = await request.json()
    
    # Обновляем поля
    if "name" in data:
        block.name = data["name"]
    if "description" in data:
        block.description = data["description"]
    
    db.commit()
    
    return JSONResponse({
        "status": "success",
        "block": {
            "id": block.id,
            "name": block.name,
            "description": block.description,
            "display_order": block.display_order
        }
    })



@app.delete("/api/blocks/{block_id}")
async def delete_block(
    block_id: int,
    current_user: User = Depends(require_role(['admin', 'controller'])),
    db: Session = Depends(get_db)
):
    """Удаление блока метрик"""
    block = db.query(MetricBlock).filter(MetricBlock.id == block_id).first()
    if not block:
        raise HTTPException(status_code=404, detail="Блок не найден")
    
    # Проверяем, есть ли метрики в этом блоке
    metrics_count = db.query(Metric).filter(Metric.block_id == block_id).count()
    if metrics_count > 0:
        # Если есть метрики, просто деактивируем блок
        block.is_active = False
        db.commit()
        return JSONResponse({
            "status": "success",
            "message": f"Блок деактивирован, так как содержит {metrics_count} метрик"
        })
    else:
        # Если нет метрик, удаляем полностью
        db.delete(block)
        db.commit()
        return JSONResponse({
            "status": "success",
            "message": "Блок удален"
        })

# ============== ПОЛЬЗОВАТЕЛИ И ПРОФИЛЬ ==============

@app.get("/profile", response_class=HTMLResponse)
async def profile_redirect():
    """Редирект на страницу профиля в auth"""
    return RedirectResponse(url="/auth/profile", status_code=302)

# ============== АДМИН ПАНЕЛЬ ==============

@app.get("/admin", response_class=HTMLResponse)
async def admin_panel(
    request: Request,
    current_user: User = Depends(require_role(['admin'])),
    db: Session = Depends(get_db)
):
    '''Админ панель'''
    users_count = db.query(User).count()
    projects_count = db.query(Project).count()
    operators_count = db.query(Operator).count()
    evaluations_count = db.query(Evaluation).count()
    
    # Количество тикетов по статусам
    tickets_new = db.query(Ticket).filter(Ticket.status == TicketStatus.NEW.value).count()
    tickets_in_progress = db.query(Ticket).filter(Ticket.status == TicketStatus.IN_PROGRESS.value).count()
    
    # Получаем все проекты для полосы проектов
    projects = db.query(Project).filter(Project.is_active == True).all()
    
    # Количество активных тикетов для бейджа
    tickets_count = get_active_tickets_count(db)
    
    return templates.TemplateResponse(
        "admin.html",
        {
            "request": request,
            "user": current_user,
            "projects": projects,
            "stats": {
                "users": users_count,
                "projects": projects_count,
                "operators": operators_count,
                "evaluations": evaluations_count,
                "tickets_new": tickets_new,
                "tickets_in_progress": tickets_in_progress
            },
            "tickets_count": tickets_count,
            "date": date
        }
    )

@app.get("/admin/users", response_class=HTMLResponse)
async def admin_users_redirect():
    """Редирект на страницу управления пользователями"""
    return RedirectResponse(url="/auth/admin/users", status_code=302)

# ============== API ДЛЯ ПРОЕКТОВ ==============

@app.get("/api/projects")
async def get_projects(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получение списка проектов"""
    projects = db.query(Project).filter(Project.is_active == True).all()
    
    result = []
    for project in projects:
        result.append({
            "id": project.id,
            "name": project.name,
            "description": project.description,
            "operators_count": db.query(Operator).filter(
                Operator.project_id == project.id,
                Operator.is_active == True
            ).count(),
            "blocks_count": db.query(MetricBlock).filter(
                MetricBlock.project_id == project.id,
                MetricBlock.is_active == True
            ).count()
        })
    
    return JSONResponse(result)

@app.get("/api/projects/{project_id}/metrics")
async def get_project_metrics(
    project_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получение метрик проекта"""
    metrics = db.query(Metric).join(MetricBlock).filter(
        MetricBlock.project_id == project_id,
        Metric.is_active == True,
        MetricBlock.is_active == True
    ).order_by(MetricBlock.display_order, Metric.display_order).all()
    
    result = []
    for metric in metrics:
        result.append({
            "id": metric.id,
            "name": metric.name,
            "description": metric.description,
            "max_score": metric.max_score,
            "is_critical": metric.is_critical,
            "is_global_critical": metric.is_global_critical,  # НОВОЕ ПОЛЕ
            "allow_na": metric.allow_na,                      # НОВОЕ ПОЛЕ
            "penalty_type": metric.penalty_type,
            "penalty_value": metric.penalty_value,
            "resets_block": metric.resets_block,
            "block_id": metric.block_id,
            "block_name": metric.block.name,
            "display_order": metric.display_order
        })
    
    return JSONResponse(result)

@app.get("/api/projects/{project_id}/blocks")
async def get_project_blocks(
    project_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получение блоков метрик проекта"""
    blocks = db.query(MetricBlock).filter(
        MetricBlock.project_id == project_id,
        MetricBlock.is_active == True
    ).order_by(MetricBlock.display_order).all()
    
    result = []
    for block in blocks:
        metrics_count = db.query(Metric).filter(
            Metric.block_id == block.id,
            Metric.is_active == True
        ).count()
        
        result.append({
            "id": block.id,
            "name": block.name,
            "description": block.description,
            "display_order": block.display_order,
            "metrics_count": metrics_count
        })
    
    return JSONResponse(result)

# ============== API ДЛЯ ОПЕРАТОРОВ ==============

@app.get("/api/operators")
async def get_operators(
    project_id: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получение списка операторов"""
    query = db.query(Operator).filter(Operator.is_active == True)
    
    if project_id:
        query = query.filter(Operator.project_id == project_id)
    
    operators = query.all()
    
    result = []
    for op in operators:
        result.append({
            "id": op.id,
            "full_name": op.full_name,
            "login": op.login,
            "avatar_letters": op.avatar_letters,
            "project_id": op.project_id,
            "project_name": op.project.name,
            "hire_date": op.hire_date.strftime("%Y-%m-%d") if op.hire_date else None,
            "experience_days": op.experience_days,
            "avg_quality": round(op.avg_quality, 1) if op.avg_quality else 0,
            "total_evaluations": op.total_evaluations
        })
    
    return JSONResponse(result)


@app.post("/api/operator/create")
async def create_operator(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Создание нового оператора"""
    if current_user.role not in ['admin', 'controller']:
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    
    try:
        data = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Неверный формат данных")
    
    # Проверяем обязательные поля
    required_fields = ['full_name', 'login', 'project_id']
    for field in required_fields:
        if not data.get(field):
            raise HTTPException(status_code=400, detail=f"Поле {field} обязательно")
    
    # Проверяем существование проекта
    project = db.query(Project).filter(Project.id == data.get("project_id")).first()
    if not project:
        raise HTTPException(status_code=400, detail="Указанный проект не существует")
    
    # Ищем существующего оператора (включая неактивных)
    existing = db.query(Operator).filter(Operator.login == data.get("login")).first()
    
    if existing:
        if existing.is_active:
            raise HTTPException(
                status_code=400, 
                detail=f"Оператор с логином '{data.get('login')}' уже существует"
            )
        else:
            # Реактивируем удаленного оператора
            existing.is_active = True
            existing.full_name = data.get("full_name")
            existing.project_id = data.get("project_id")
            
            # Обновляем дату выхода
            if data.get("hire_date"):
                try:
                    existing.hire_date = datetime.strptime(data.get("hire_date"), "%Y-%m-%d").date()
                    existing.experience_days = (date.today() - existing.hire_date).days
                except ValueError:
                    raise HTTPException(status_code=400, detail="Неверный формат даты")
            
            # Обновляем аватар
            full_name = data.get("full_name", "")
            if full_name:
                name_parts = full_name.split()
                if len(name_parts) >= 2:
                    existing.avatar_letters = name_parts[0][0] + name_parts[1][0]
                else:
                    existing.avatar_letters = full_name[:2].upper()
            
            try:
                db.commit()
                db.refresh(existing)
                
                return JSONResponse({
                    "status": "success",
                    "operator": {
                        "id": existing.id,
                        "full_name": existing.full_name,
                        "login": existing.login,
                        "avatar_letters": existing.avatar_letters,
                        "project_id": existing.project_id,
                        "project_name": project.name,
                        "hire_date": existing.hire_date.strftime("%Y-%m-%d") if existing.hire_date else None,
                        "experience_days": existing.experience_days,
                        "is_active": existing.is_active
                    },
                    "message": "Оператор восстановлен"
                })
            except Exception as e:
                db.rollback()
                raise HTTPException(status_code=500, detail=f"Ошибка при восстановлении: {str(e)}")
    
    # Если оператор не найден - создаем нового
    # Генерируем инициалы
    full_name = data.get("full_name", "")
    avatar_letters = None
    if full_name:
        name_parts = full_name.split()
        if len(name_parts) >= 2:
            avatar_letters = name_parts[0][0] + name_parts[1][0]
        else:
            avatar_letters = full_name[:2].upper()
    
    # Рассчитываем стаж
    hire_date = None
    experience_days = 0
    if data.get("hire_date"):
        try:
            hire_date = datetime.strptime(data.get("hire_date"), "%Y-%m-%d").date()
            experience_days = (date.today() - hire_date).days
        except ValueError:
            raise HTTPException(status_code=400, detail="Неверный формат даты")
    
    # Создаем оператора
    operator = Operator(
        full_name=full_name,
        login=data.get("login"),
        avatar_letters=avatar_letters,
        hire_date=hire_date,
        experience_days=experience_days,
        project_id=data.get("project_id"),
        is_active=True
    )
    
    try:
        db.add(operator)
        db.commit()
        db.refresh(operator)
        
        return JSONResponse({
            "status": "success",
            "operator": {
                "id": operator.id,
                "full_name": operator.full_name,
                "login": operator.login,
                "avatar_letters": operator.avatar_letters,
                "project_id": operator.project_id,
                "project_name": project.name,
                "hire_date": operator.hire_date.strftime("%Y-%m-%d") if operator.hire_date else None,
                "experience_days": operator.experience_days,
                "is_active": operator.is_active
            }
        })
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка при сохранении: {str(e)}")

@app.put("/api/operator/{operator_id}")
async def update_operator(
    operator_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Обновление данных оператора"""
    if current_user.role not in ['admin', 'controller']:
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    
    operator = db.query(Operator).filter(Operator.id == operator_id).first()
    if not operator:
        raise HTTPException(status_code=404, detail="Оператор не найден")
    
    data = await request.json()
    
    # Обновляем поля
    if "full_name" in data and data["full_name"]:
        operator.full_name = data["full_name"]
        # Обновляем инициалы
        name_parts = data["full_name"].split()
        if len(name_parts) >= 2:
            operator.avatar_letters = name_parts[0][0] + name_parts[1][0]
        else:
            operator.avatar_letters = data["full_name"][:2].upper()
    
    if "login" in data and data["login"]:
        # Проверяем уникальность логина
        existing = db.query(Operator).filter(
            Operator.login == data["login"],
            Operator.id != operator_id
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="Логин уже используется")
        operator.login = data["login"]
    
    if "hire_date" in data:
        if data["hire_date"]:
            operator.hire_date = datetime.strptime(data["hire_date"], "%Y-%m-%d").date()
            operator.experience_days = (date.today() - operator.hire_date).days
        else:
            operator.hire_date = None
            operator.experience_days = 0
    
    if "project_id" in data:
        operator.project_id = data["project_id"]
    
    if "is_active" in data:
        operator.is_active = data["is_active"]
    
    db.commit()
    db.refresh(operator)
    
    return JSONResponse({
        "status": "success",
        "operator": {
            "id": operator.id,
            "full_name": operator.full_name,
            "login": operator.login,
            "avatar_letters": operator.avatar_letters,
            "project_id": operator.project_id,
            "hire_date": operator.hire_date.strftime("%Y-%m-%d") if operator.hire_date else None,
            "experience_days": operator.experience_days,
            "is_active": operator.is_active
        }
    })


@app.delete("/api/operator/{operator_id}")
async def delete_operator(
    operator_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Удаление оператора (мягкое удаление - деактивация)"""
    if current_user.role not in ['admin', 'controller']:
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    
    operator = db.query(Operator).filter(Operator.id == operator_id).first()
    if not operator:
        raise HTTPException(status_code=404, detail="Оператор не найден")
    
    # Мягкое удаление - просто деактивируем
    operator.is_active = False
    db.commit()
    
    return JSONResponse({
        "status": "success",
        "message": f"Оператор {operator.full_name} деактивирован"
    })

# ============== УПРАВЛЕНИЕ ТЕМАМИ ЗВОНКОВ ==============

@app.get("/project/{project_id}/topics", response_class=HTMLResponse)
async def project_topics_page(
    request: Request,
    project_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Страница управления темами звонков проекта"""
    if current_user.role not in ['admin', 'controller']:
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Проект не найден")
    
    # Получаем темы проекта
    topics = db.query(Topic).filter(
        Topic.project_id == project_id,
        Topic.is_active == True
    ).order_by(Topic.display_order).all()
    
    # Получаем все проекты для полосы проектов
    all_projects = db.query(Project).filter(Project.is_active == True).all()
    
    # Количество активных тикетов для бейджа
    tickets_count = get_active_tickets_count(db)
    
    return templates.TemplateResponse(
        "project_topics.html",
        {
            "request": request,
            "user": current_user,
            "project": project,
            "topics": topics,
            "projects": all_projects,
            "tickets_count": tickets_count,
            "date": date
        }
    )


@app.post("/api/project/{project_id}/topics")
async def create_topic(
    project_id: int,
    request: Request,
    current_user: User = Depends(require_role(['admin', 'controller'])),
    db: Session = Depends(get_db)
):
    """Создание новой темы звонка"""
    data = await request.json()
    
    # Проверяем существование проекта
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Проект не найден")
    
    # Получаем максимальный порядок отображения
    max_order = db.query(func.max(Topic.display_order)).filter(
        Topic.project_id == project_id
    ).scalar() or 0
    
    topic = Topic(
        project_id=project_id,
        name=data.get("name"),
        description=data.get("description", ""),
        display_order=max_order + 1,
        is_active=True
    )
    db.add(topic)
    db.commit()
    db.refresh(topic)
    
    return JSONResponse({
        "status": "success",
        "topic": {
            "id": topic.id,
            "name": topic.name,
            "description": topic.description,
            "display_order": topic.display_order
        }
    })


@app.put("/api/topics/{topic_id}")
async def update_topic(
    topic_id: int,
    request: Request,
    current_user: User = Depends(require_role(['admin', 'controller'])),
    db: Session = Depends(get_db)
):
    """Обновление темы звонка"""
    topic = db.query(Topic).filter(Topic.id == topic_id).first()
    if not topic:
        raise HTTPException(status_code=404, detail="Тема не найдена")
    
    data = await request.json()
    
    topic.name = data.get("name", topic.name)
    topic.description = data.get("description", topic.description)
    topic.is_active = data.get("is_active", topic.is_active)
    
    db.commit()
    
    return JSONResponse({"status": "success"})


@app.delete("/api/topics/{topic_id}")
async def delete_topic(
    topic_id: int,
    current_user: User = Depends(require_role(['admin', 'controller'])),
    db: Session = Depends(get_db)
):
    """Удаление темы звонка"""
    topic = db.query(Topic).filter(Topic.id == topic_id).first()
    if not topic:
        raise HTTPException(status_code=404, detail="Тема не найдена")
    
    # Проверяем, есть ли оценки с этой темой
    evaluations_count = db.query(Evaluation).filter(Evaluation.topic_id == topic_id).count()
    if evaluations_count > 0:
        # Если есть оценки, просто деактивируем
        topic.is_active = False
        db.commit()
        return JSONResponse({
            "status": "success",
            "message": f"Тема деактивирована, так как используется в {evaluations_count} оценках"
        })
    else:
        # Если нет оценок, удаляем полностью
        db.delete(topic)
        db.commit()
        return JSONResponse({"status": "success", "message": "Тема удалена"})


@app.post("/api/projects/{project_id}/topics/reorder")
async def reorder_topics(
    project_id: int,
    request: Request,
    current_user: User = Depends(require_role(['admin', 'controller'])),
    db: Session = Depends(get_db)
):
    """Изменение порядка тем"""
    data = await request.json()
    orders = data.get("orders", [])
    
    for item in orders:
        topic = db.query(Topic).filter(
            Topic.id == item["id"],
            Topic.project_id == project_id
        ).first()
        if topic:
            topic.display_order = item["order"]
    
    db.commit()
    return JSONResponse({"status": "success"})

# ============== ПОИСК ==============

@app.get("/api/search")
async def global_search(
    q: str = Query(..., min_length=2),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Глобальный поиск по операторам и проектам"""
    results = {
        "operators": [],
        "projects": []
    }
    
    operators = db.query(Operator).filter(
        Operator.is_active == True,
        Operator.full_name.ilike(f"%{q}%")
    ).limit(5).all()
    
    for op in operators:
        results["operators"].append({
            "id": op.id,
            "full_name": op.full_name,
            "login": op.login,
            "project": op.project.name,
            "url": f"/operator/{op.id}"
        })
    
    projects = db.query(Project).filter(
        Project.is_active == True,
        Project.name.ilike(f"%{q}%")
    ).limit(5).all()
    
    for proj in projects:
        results["projects"].append({
            "id": proj.id,
            "name": proj.name,
            "url": f"/project/{proj.id}"
        })
    
    return JSONResponse(results)


# ============== МЕТРИКИ ОПЕРАТОРА ДЛЯ ФИЛЬТРА ==============

@app.get("/api/operator/{operator_id}/metrics")
async def get_operator_metrics(
    operator_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Получение списка метрик для фильтра"""
    operator = db.query(Operator).filter(Operator.id == operator_id).first()
    if not operator:
        raise HTTPException(status_code=404, detail="Оператор не найден")
    
    # Получаем метрики через блоки
    metrics = db.query(Metric).join(MetricBlock).filter(
        MetricBlock.project_id == operator.project_id,
        Metric.is_active == True,
        MetricBlock.is_active == True
    ).order_by(MetricBlock.display_order, Metric.display_order).all()
    
    result = []
    for metric in metrics:
        result.append({
            "id": metric.id,
            "name": metric.name,
            "max_score": metric.max_score,
            "block_name": metric.block.name
        })
    
    return JSONResponse(result)


# ============== ЗАПУСК ==============

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)